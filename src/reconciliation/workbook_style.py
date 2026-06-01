"""Centralised, subtle professional styling for reconciliation workbooks.

This module holds the visual vocabulary (palette, fonts, borders, number
formats) and small deterministic helpers used by the workbook writer. Styling
is intentionally conservative: dark/medium headers, light section fills, thin
borders, frozen panes and filters. Nothing here changes a single financial
value; it only affects presentation.
"""

from __future__ import annotations

from typing import Iterable

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet

# --- Palette (ARGB hex, no leading '#') -------------------------------------
# Dark blue/grey for main headers, light blue/grey for section headers, and a
# small set of status fills used sparingly and only where they aid review.
COLOR_HEADER_DARK = "1F3864"      # dark navy - main table headers
COLOR_HEADER_MED = "2F5496"       # medium blue - secondary headers
COLOR_SECTION = "D9E1F2"          # light blue/grey - section bands
COLOR_BANNER = "44546A"           # slate - title banners
COLOR_REVIEW = "FFF2CC"           # light yellow - review-required areas
COLOR_MATCH = "E2EFDA"            # light green - strong matched indicators
COLOR_FAIL = "F8CBAD"             # light red - validation fail / critical
COLOR_ZEBRA = "F2F2F2"            # very light grey - optional banding
COLOR_WHITE = "FFFFFF"

# --- Number formats ---------------------------------------------------------
FMT_AMOUNT = "#,##0.00;[Red](#,##0.00)"
FMT_DATE = "yyyy-mm-dd"

# --- Reusable style objects -------------------------------------------------
FONT_TITLE = Font(name="Calibri", size=14, bold=True, color=COLOR_WHITE)
FONT_HEADER = Font(name="Calibri", size=10, bold=True, color=COLOR_WHITE)
FONT_SECTION = Font(name="Calibri", size=10, bold=True, color=COLOR_HEADER_DARK)
FONT_LABEL = Font(name="Calibri", size=10, bold=True)
FONT_NORMAL = Font(name="Calibri", size=10)
FONT_NOTE = Font(name="Calibri", size=9, italic=True, color="595959")

FILL_HEADER = PatternFill("solid", fgColor=COLOR_HEADER_DARK)
FILL_HEADER_MED = PatternFill("solid", fgColor=COLOR_HEADER_MED)
FILL_SECTION = PatternFill("solid", fgColor=COLOR_SECTION)
FILL_BANNER = PatternFill("solid", fgColor=COLOR_BANNER)
FILL_REVIEW = PatternFill("solid", fgColor=COLOR_REVIEW)
FILL_MATCH = PatternFill("solid", fgColor=COLOR_MATCH)
FILL_FAIL = PatternFill("solid", fgColor=COLOR_FAIL)

_THIN = Side(style="thin", color="BFBFBF")
BORDER_THIN = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
ALIGN_LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

# Column header names that should carry the amount number format wherever they
# appear. Matched by exact header text.
AMOUNT_HEADERS = frozenset(
    {
        "debit_source",
        "credit_source",
        "net_source",
        "debit_org_perspective",
        "credit_org_perspective",
        "net_org_perspective",
        "org_amount",
        "party_amount",
        "amount_difference",
        "amount",
        "amount_org_perspective",
        "Org Ledger Amount",
        "Party Ledger Amount",
        "Difference",
        "Amount (Rs.)",
    }
)

DATE_HEADERS = frozenset({"date", "org_date", "party_date", "normalized_date", "Date"})


def style_table_header(ws: Worksheet, row: int, n_cols: int, *, medium: bool = False) -> None:
    """Apply a bold filled header band with thin borders to a header row."""
    fill = FILL_HEADER_MED if medium else FILL_HEADER
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = FONT_HEADER
        cell.fill = fill
        cell.border = BORDER_THIN
        cell.alignment = ALIGN_CENTER


def style_section_title(ws: Worksheet, row: int, *, span: int = 1) -> None:
    """Style a section banner row (light blue band, bold dark text)."""
    for c in range(1, max(span, 1) + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_SECTION
        cell.font = FONT_SECTION
        cell.border = BORDER_THIN
        cell.alignment = ALIGN_LEFT


def style_banner(ws: Worksheet, row: int, span: int) -> None:
    """Style a top title banner (slate fill, white bold)."""
    for c in range(1, max(span, 1) + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_BANNER
        cell.font = FONT_TITLE
        cell.alignment = ALIGN_LEFT


def apply_borders(ws: Worksheet, first_row: int, last_row: int, n_cols: int) -> None:
    """Thin-border a rectangular block (header + data)."""
    if last_row < first_row:
        return
    for r in range(first_row, last_row + 1):
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).border = BORDER_THIN


def apply_number_formats(
    ws: Worksheet,
    headers: list[str],
    first_data_row: int,
    last_row: int,
) -> None:
    """Apply amount/date number formats by header name across a data block."""
    if last_row < first_data_row:
        return
    for idx, header in enumerate(headers, start=1):
        name = str(header)
        if name in AMOUNT_HEADERS:
            fmt = FMT_AMOUNT
        elif name in DATE_HEADERS:
            fmt = FMT_DATE
        else:
            continue
        for r in range(first_data_row, last_row + 1):
            ws.cell(row=r, column=idx).number_format = fmt


def autosize(ws: Worksheet, max_col: int, max_row: int, *, cap: int = 58) -> None:
    """Width columns to content with a sensible cap; scans a bounded window."""
    for c in range(1, max_col + 1):
        letter = get_column_letter(c)
        longest = 0
        for r in range(1, min(max_row, 400) + 1):
            value = ws.cell(r, c).value
            if value is None:
                continue
            text = str(value)
            # Don't let long formulas blow out the width.
            if text.startswith("="):
                text = text[:18]
            longest = max(longest, len(text))
        ws.column_dimensions[letter].width = max(10, min(longest + 2, cap))


def status_fill(match_status: str, review_required: bool) -> PatternFill | None:
    """Subtle row fill by reconciliation status (used sparingly)."""
    status = (match_status or "").strip().lower()
    if status in {"matched_strong", "matched_supported"}:
        return FILL_MATCH
    if review_required or status == "candidate_review":
        return FILL_REVIEW
    if status in {"unmatched_org", "unmatched_party"}:
        return FILL_FAIL
    return None


def freeze(ws: Worksheet, cell: str) -> None:
    ws.freeze_panes = cell


def set_filter(ws: Worksheet, ref: str) -> None:
    ws.auto_filter.ref = ref


def zebra(ws: Worksheet, first_data_row: int, last_row: int, n_cols: int) -> None:
    """Optional very-light banding to improve readability of long tables."""
    fill = PatternFill("solid", fgColor=COLOR_ZEBRA)
    for r in range(first_data_row, last_row + 1):
        if (r - first_data_row) % 2 == 1:
            for c in range(1, n_cols + 1):
                cell = ws.cell(row=r, column=c)
                if cell.fill is None or cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = fill


def style_key_value_block(ws: Worksheet, rows: Iterable[int], *, key_col: int = 1) -> None:
    """Bold the key column of an identity / key-value block."""
    for r in rows:
        ws.cell(row=r, column=key_col).font = FONT_LABEL
        ws.cell(row=r, column=key_col).alignment = ALIGN_LEFT


# --- Reference-derived theming + print/page setup ---------------------------
# Module-level page state applied by ``apply_page_setup``. Defaults suit a
# wide finance reconciliation workbook (landscape, fit columns to one page wide).
_PAGE_ORIENTATION = "landscape"
_FIT_TO_WIDTH = 1


def apply_theme(theme) -> None:
    """Rebind the shared style objects from a derived ``WorkbookTheme``.

    Called once before writing so the workbook adopts the reference workbook's
    header/banner palette and page orientation while preserving the clean,
    audit-friendly structure. Invariants are guaranteed: header fonts stay bold
    and the amount format keeps a thousands separator, so formula/format audits
    and presentation tests remain valid.
    """
    global FONT_TITLE, FONT_HEADER, FILL_HEADER, FILL_HEADER_MED, FILL_BANNER
    global FILL_SECTION, FMT_AMOUNT, FMT_DATE, _PAGE_ORIENTATION, _FIT_TO_WIDTH

    header_fill = getattr(theme, "header_fill", COLOR_HEADER_DARK)
    header_font_color = getattr(theme, "header_font_color", COLOR_WHITE) or COLOR_WHITE
    banner_fill = getattr(theme, "banner_fill", COLOR_BANNER)
    section_fill = getattr(theme, "section_fill", COLOR_SECTION)
    title_font_name = getattr(theme, "title_font_name", "Calibri") or "Calibri"

    FILL_HEADER = PatternFill("solid", fgColor=header_fill)
    FILL_HEADER_MED = PatternFill("solid", fgColor=header_fill)
    FILL_BANNER = PatternFill("solid", fgColor=banner_fill)
    FILL_SECTION = PatternFill("solid", fgColor=section_fill)
    FONT_TITLE = Font(name=title_font_name, size=14, bold=True, color=COLOR_WHITE)
    FONT_HEADER = Font(name=title_font_name, size=10, bold=True, color=header_font_color)

    amount_fmt = getattr(theme, "amount_format", FMT_AMOUNT) or FMT_AMOUNT
    if "#,##0" in amount_fmt:  # guard the audited invariant
        FMT_AMOUNT = amount_fmt
    FMT_DATE = getattr(theme, "date_format", FMT_DATE) or FMT_DATE

    _PAGE_ORIENTATION = getattr(theme, "page_orientation", _PAGE_ORIENTATION) or _PAGE_ORIENTATION
    _FIT_TO_WIDTH = int(getattr(theme, "fit_to_width", _FIT_TO_WIDTH) or _FIT_TO_WIDTH)


def apply_page_setup(ws: Worksheet, *, repeat_header_row: int | None = 1) -> None:
    """Apply print orientation, fit-to-width and repeated header rows to a sheet."""
    ws.page_setup.orientation = _PAGE_ORIENTATION
    ws.page_setup.fitToWidth = _FIT_TO_WIDTH
    ws.page_setup.fitToHeight = 0
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    else:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = False
    if repeat_header_row:
        ws.print_title_rows = f"{repeat_header_row}:{repeat_header_row}"
