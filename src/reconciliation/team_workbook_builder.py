"""Build a clean, team-facing workbook from the internal reconciliation output."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from src.reconciliation import workbook_style as st

TEAM_REVIEW_COLUMNS = [
    "priority",
    "source_sheet",
    "source_row_id",
    "type_label",
    "date",
    "reference",
    "amount",
    "match_status",
    "reason",
    "suggested_action",
    "reviewer_comment",
    "manual_status",
    "resolved_by",
    "resolved_date",
]

INTERNAL_SHEETS = {
    "README",
    "Executive_Summary",
    "AI_Decision_Audit",
    "Match_Evidence",
    "Master_Match_Table",
    "Review_Queue",
    "Validation_Report",
    "Formula_Audit",
    "Assumptions_And_Limits",
}

WORKING_COLUMNS_TO_HIDE = {
    "match_key_primary",
    "match_key_secondary",
    "decision_source",
    "match_type",
    "match_confidence",
    "validation_status",
    "raw_text",
}

VISIBLE_STATUS = {
    "matched_strong": "Matched",
    "matched_supported": "Matched",
    "matched_ai": "Matched",
    "ledger_only_ai": "Ledger only",
    "candidate_review": "Needs review",
    "unmatched_org": "Needs review",
    "unmatched_party": "Needs review",
}


def _headers(ws, row: int = 1) -> list[str]:
    return [str(cell.value or "") for cell in ws[row]]


def _table_name(seed: str, used: set[str]) -> str:
    base = re.sub(r"[^A-Za-z0-9_]", "_", seed).strip("_") or "TeamTable"
    if base[0].isdigit():
        base = f"T_{base}"
    name = base[:240]
    suffix = 2
    while name.lower() in used:
        tail = f"_{suffix}"
        name = f"{base[:240 - len(tail)]}{tail}"
        suffix += 1
    used.add(name.lower())
    return name


def _add_excel_table(ws, name: str, used: set[str]) -> None:
    if ws.max_row < 2 or ws.max_column < 1:
        return
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName=_table_name(name, used), ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _plain_reason(value: object) -> str:
    text = str(value or "").strip()
    lower = text.lower()
    if "ai_" in lower or " ai " in f" {lower} " or "provider" in lower or "token" in lower:
        return "Automated reconciliation could not confidently place this row. Review against both ledgers."
    return text


def _plain_action(value: object) -> str:
    text = str(value or "").strip()
    if "ai" in text.lower():
        return "Review this row against both source ledgers and confirm the correct treatment."
    return text


def _write_team_guide(ws, *, pair_id: str) -> None:
    ws["A1"] = "TEAM RECONCILIATION WORKBOOK"
    ws.merge_cells("A1:D1")
    st.style_banner(ws, 1, 4)
    rows = [
        ("Pair ID", pair_id),
        ("Purpose", "Clean finance-team view of the standardized ledgers, summary, annexures, and rows needing improvement."),
        ("Start here", "Review Summary, then Rows_Needing_Improvement, then the relevant annexure."),
        ("Ledger entry", "Paste new rows directly below the last row of an intended Working ledger table. Excel expands the table and calculated amount columns."),
        ("Refresh rule", "Rerun the reconciliation engine after adding source rows so match status, annexures, and the improvement list are regenerated authoritatively."),
        ("Hidden detail", "Internal audit evidence remains embedded but hidden from the normal team workflow."),
    ]
    for row_no, (caption, value) in enumerate(rows, start=3):
        ws.cell(row=row_no, column=1, value=caption).font = st.FONT_LABEL
        ws.cell(row=row_no, column=2, value=value)
        ws.merge_cells(start_row=row_no, start_column=2, end_row=row_no, end_column=4)
        ws.cell(row=row_no, column=2).alignment = st.ALIGN_LEFT_WRAP
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    st.freeze(ws, "A3")


def _write_improvement_rows(wb) -> None:
    source = wb["Review_Queue"]
    source_headers = _headers(source)
    source_idx = {name: idx for idx, name in enumerate(source_headers)}
    ws = wb.create_sheet("Rows_Needing_Improvement")
    for col_no, name in enumerate(TEAM_REVIEW_COLUMNS, start=1):
        ws.cell(row=1, column=col_no, value=name)
    for row_no, source_row in enumerate(source.iter_rows(min_row=2, values_only=True), start=2):
        payload: dict[str, Any] = {
            name: source_row[source_idx[name]] if name in source_idx else ""
            for name in TEAM_REVIEW_COLUMNS
        }
        payload["reason"] = _plain_reason(payload["reason"])
        payload["suggested_action"] = _plain_action(payload["suggested_action"])
        payload["match_status"] = VISIBLE_STATUS.get(
            str(payload["match_status"] or ""), payload["match_status"]
        )
        if payload["source_sheet"] == "Match_Evidence":
            payload["source_sheet"] = "Reconciliation"
        for col_no, name in enumerate(TEAM_REVIEW_COLUMNS, start=1):
            ws.cell(row=row_no, column=col_no, value=payload[name])
    st.style_table_header(ws, 1, len(TEAM_REVIEW_COLUMNS))
    if ws.max_row >= 2:
        st.apply_borders(ws, 1, ws.max_row, len(TEAM_REVIEW_COLUMNS))
        st.apply_number_formats(ws, TEAM_REVIEW_COLUMNS, 2, ws.max_row)
        for row_no in range(2, ws.max_row + 1):
            ws.cell(row=row_no, column=1).fill = st.FILL_REVIEW
    st.freeze(ws, "A2")
    st.set_filter(ws, f"A1:{get_column_letter(len(TEAM_REVIEW_COLUMNS))}{max(1, ws.max_row)}")
    st.autosize(ws, len(TEAM_REVIEW_COLUMNS), max(1, ws.max_row))


def _hide_working_technical_columns(ws) -> None:
    for idx, name in enumerate(_headers(ws), start=1):
        if name in WORKING_COLUMNS_TO_HIDE:
            ws.column_dimensions[get_column_letter(idx)].hidden = True
        if name in {"decision_source", "match_type"}:
            for row_no in range(2, ws.max_row + 1):
                ws.cell(row=row_no, column=idx, value="")


def _sanitize_visible_status_values(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if cell.value in VISIBLE_STATUS:
                cell.value = VISIBLE_STATUS[cell.value]


def _simplify_annexure_match_section(ws) -> None:
    section_row = None
    for row in ws.iter_rows():
        if any(cell.value == "Section D: Match / Candidate view" for cell in row):
            section_row = row[0].row
            break
    if section_row is None:
        return
    header_row = section_row + 1
    headers = _headers(ws, header_row)
    mapping = {
        "match_rule": "reconciliation_basis",
        "match_confidence": "confidence_band",
        "decision_source": "placement_status",
        "match_type": "match_category",
    }
    indices = {name: idx + 1 for idx, name in enumerate(headers)}
    for original, replacement in mapping.items():
        if original in indices:
            ws.cell(row=header_row, column=indices[original], value=replacement)
    status_col = indices.get("match_status")
    for row_no in range(header_row + 1, ws.max_row + 1):
        status = str(ws.cell(row=row_no, column=status_col).value or "") if status_col else ""
        accepted = status in {"matched_strong", "matched_supported", "matched_ai"}
        basis = {
            "matched_strong": "Reference-backed match",
            "matched_supported": "Unique mirrored amount/date/type match",
            "matched_ai": "Accepted after additional validation",
        }.get(status, "Review required")
        if "match_rule" in indices:
            ws.cell(row=row_no, column=indices["match_rule"], value=basis)
        if "match_confidence" in indices:
            ws.cell(row=row_no, column=indices["match_confidence"], value="High" if accepted else "Review")
        if "decision_source" in indices:
            ws.cell(row=row_no, column=indices["decision_source"], value="Accepted" if accepted else "Review required")
        if "match_type" in indices:
            ws.cell(row=row_no, column=indices["match_type"], value="Matched" if accepted else "Needs review")
    _sanitize_visible_status_values(ws)


def build_team_workbook(
    *,
    internal_path: Path,
    output_path: Path,
    pair_id: str,
    raw_sheet_names: list[str],
) -> Path:
    """Save a clean workbook while preserving hidden formula dependencies."""
    wb = load_workbook(internal_path)
    guide = wb.create_sheet("Team_Guide", 0)
    _write_team_guide(guide, pair_id=pair_id)
    _write_improvement_rows(wb)

    used_table_names: set[str] = set()
    for ws in wb.worksheets:
        if ws.title.startswith("Working "):
            _hide_working_technical_columns(ws)
            _sanitize_visible_status_values(ws)
            _add_excel_table(ws, f"Standardized_{ws.title}", used_table_names)
        elif ws.title.startswith("Annex_"):
            _simplify_annexure_match_section(ws)
    _add_excel_table(wb["Rows_Needing_Improvement"], "Rows_Needing_Improvement", used_table_names)

    hidden = INTERNAL_SHEETS | set(raw_sheet_names)
    for name in hidden:
        if name in wb.sheetnames:
            wb[name].sheet_state = "hidden"

    visible_order = (
        ["Team_Guide"]
        + [name for name in wb.sheetnames if name.startswith("Working ")]
        + ["Summary"]
        + [name for name in wb.sheetnames if name.startswith("Annex_")]
        + (["Unknown_Needs_Review"] if "Unknown_Needs_Review" in wb.sheetnames else [])
        + ["Rows_Needing_Improvement"]
    )
    remaining = [name for name in wb.sheetnames if name not in visible_order]
    wb._sheets = [wb[name] for name in visible_order + remaining]
    if hasattr(wb, "calculation"):
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(output_path)
    except PermissionError:
        fallback = output_path.with_name(
            f"{output_path.stem}__locked_fallback{output_path.suffix}"
        )
        wb.save(fallback)
        wb.close()
        return fallback
    wb.close()
    return output_path
