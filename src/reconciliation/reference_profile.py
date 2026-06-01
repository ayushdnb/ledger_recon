"""Reference workbook discovery and profiling helpers."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.config import settings


# A reference workbook whose name resembles this is strongly preferred when
# several candidate workbooks are discovered.
PREFERRED_REFERENCE_HINT = "startling achievement"


def _looks_like_reference(path: Path) -> bool:
    name = path.name.lower()
    if name.startswith("~$"):
        return False
    # Never treat our own generated artefacts as the reference workbook.
    skip_tokens = (
        "formalized_ledgers__",
        "recon_workbook__",
        "raw_pdf_context__",
        "reference_workbook_profile__",
    )
    return not any(token in name for token in skip_tokens)


def discover_reference_workbook(explicit_path: str | None = None) -> Path | None:
    """Find the reference workbook by searching several locations recursively.

    Order of search roots (first match wins, preferring a name that resembles
    the known reference workbook):

    1. an explicit path, if provided and existing
    2. ``data/03_reference_workbooks`` (recursive)
    3. the project root (recursive)
    4. ``data`` (recursive)

    The original (source) reference workbook is never mutated; it is only read.
    """
    if explicit_path:
        path = Path(explicit_path)
        return path if path.exists() else None

    root = settings.project_root
    search_roots = [
        root / "data" / "03_reference_workbooks",
        root,
        root / "data",
    ]

    ordered: list[Path] = []
    seen: set[Path] = set()
    for base in search_roots:
        if not base.exists():
            continue
        for p in sorted(base.glob("**/*.xlsx")):
            resolved = p.resolve()
            if resolved in seen or not _looks_like_reference(p):
                continue
            seen.add(resolved)
            ordered.append(p)

    if not ordered:
        return None
    preferred = [p for p in ordered if PREFERRED_REFERENCE_HINT in p.name.lower()]
    return preferred[0] if preferred else ordered[0]


def profile_reference_workbook(path: Path) -> dict[str, Any]:
    """Read lightweight structural profile from a real workbook file."""
    wb = load_workbook(path, data_only=False, read_only=False)
    sheets: list[dict[str, Any]] = []
    for name in wb.sheetnames:
        ws = wb[name]
        headers = [c.value for c in ws[1]]
        formula_count = 0
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 500)):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
        sheets.append(
            {
                "sheet_name": name,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "headers": [h for h in headers if h is not None],
                "formula_count": formula_count,
                "freeze_panes": str(ws.freeze_panes or ""),
            }
        )
    wb.close()
    return {
        "workbook_name": path.name,
        "workbook_path": str(path),
        "sheet_count": len(wb.sheetnames),
        "sheet_names": wb.sheetnames,
        "sheets": sheets,
    }


def load_existing_reference_profile() -> dict[str, Any] | None:
    """Load precomputed reference profile JSON if available."""
    profile_dir = settings.project_root / "data/04_outputs/reference_profiles"
    if not profile_dir.exists():
        return None
    files = sorted(profile_dir.glob("*.json"))
    if not files:
        return None
    return json.loads(files[0].read_text(encoding="utf-8"))

