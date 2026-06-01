"""Write a submission-quality reconciliation workbook.

The workbook is formula-backed. Computed amounts are Excel formulas (so a
reviewer can audit them live), evidence/source columns are never overwritten by
a formula, and bounded AI decisions remain visible and deterministically
post-validated.

Styling is centralised in :mod:`workbook_style` and is intentionally subtle.
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from src.reconciliation import workbook_style as st
from src.reconciliation.annexure_builder import (
    ANNEX_MATCH_HEADERS,
    ANNEX_ORG_HEADERS,
    ANNEX_PARTY_HEADERS,
    AnnexurePlan,
)
from src.reconciliation.formula_builder import (
    audit_formula_cells,
    audit_protected_cells,
)
from src.reconciliation.recon_models import FormulaAuditItem, MatchRecord, ReconRow, ValidationItem
from src.reconciliation.review_queue_builder import (
    REVIEW_QUEUE_COLUMNS,
    build_review_queue,
)
from src.reconciliation.summary_builder import (
    COL_ANNEX,
    COL_DIFF,
    COL_ORG,
    COL_PARTICULARS,
    COL_PARTY,
    COL_REMARKS,
    COL_REVIEW,
    COL_REVIEWER,
    SummaryLayout,
)
from src.reconciliation.unknown_label_builder import (
    UNKNOWN_GROUPING_COLUMNS,
    build_unknown_label_groups,
    has_unknown_rows,
    unknown_rows_for_side,
)
from src.reconciliation.working_ledger_builder import WORKING_COLUMNS, WorkingRow

# Dedicated review annexure for unclassified transactions. Deliberately NOT
# prefixed "Annex_" so it never pollutes the canonical label annexure set.
UNKNOWN_REVIEW_SHEET = "Unknown_Needs_Review"

# Evidence subset shown for each unknown transaction (source-true, no formulas).
UNKNOWN_TX_HEADERS = [
    "row_id",
    "page_number",
    "date",
    "raw_type",
    "particulars",
    "reference_no",
    "debit_source",
    "credit_source",
    "review_reason",
    "ai_label_suggested",
]

# Evidence/source columns that must never be overwritten by a formula.
PROTECTED_MATCH_COLUMNS = [
    "org_raw_type",
    "party_raw_type",
    "org_particulars",
    "party_particulars",
    "org_reference",
    "party_reference",
    "org_normalized_reference",
    "party_normalized_reference",
]

MATCH_COLUMNS = [
    "match_group_id",
    "match_status",
    "match_confidence",
    "match_rule",
    "type_label",
    "org_row_id",
    "party_row_id",
    "org_date",
    "party_date",
    "date_delta_days",
    "org_reference",
    "party_reference",
    "org_normalized_reference",
    "party_normalized_reference",
    "reference_relation",
    "org_amount",
    "party_amount",
    "amount_difference",
    "org_raw_type",
    "party_raw_type",
    "org_particulars",
    "party_particulars",
    "review_required",
    "review_reason",
    "reviewer_comment",
    "manual_status",
    "decision_id",
    "decision_source",
    "match_type",
    "org_row_ids",
    "party_row_ids",
    "compact_explanation",
    "validation_status",
    "validation_reason",
    "prompt_fingerprint",
    "response_fingerprint",
    "cache_key",
]

AI_AUDIT_COLUMNS = [
    "decision_id",
    "decision_source",
    "match_status",
    "match_type",
    "match_confidence",
    "org_row_ids",
    "party_row_ids",
    "org_amount",
    "party_amount",
    "amount_difference",
    "date_delta_days",
    "match_rule",
    "compact_explanation",
    "validation_status",
    "validation_reason",
    "prompt_fingerprint",
    "response_fingerprint",
    "cache_key",
]

MATCH_EVIDENCE_SHEET = "Match_Evidence"
AI_DECISION_AUDIT_SHEET = "AI_Decision_Audit"
REQUIRED_SHEET_ORDER_HEAD = ["README", "Executive_Summary"]

# Status column letters in dependent sheets (kept here so Summary / Executive
# formulas can reference them without re-deriving from headers each time).
_VALIDATION_STATUS_COL = "C"   # check, scope, status, count, details
_FORMULA_AUDIT_STATUS_COL = "F"  # sheet_name, cell, category, present, text, status, issue


def _col_letter(columns: list[str], name: str) -> str:
    return get_column_letter(columns.index(name) + 1)


def _finalize_table(ws, header_row: int, n_cols: int, last_row: int, headers: list[str]) -> None:
    """Apply header band, borders, number formats, freeze and filter to a table."""
    st.style_table_header(ws, header_row, n_cols)
    if last_row >= header_row + 1:
        st.apply_borders(ws, header_row, last_row, n_cols)
        st.apply_number_formats(ws, headers, header_row + 1, last_row)
    st.freeze(ws, f"A{header_row + 1}")
    st.set_filter(ws, f"A{header_row}:{get_column_letter(n_cols)}{max(last_row, header_row)}")
    st.autosize(ws, n_cols, max(last_row, header_row))


def _write_rows(ws, start_row: int, columns: list[str], rows: list[dict[str, Any]]) -> int:
    for idx, name in enumerate(columns, start=1):
        ws.cell(row=start_row, column=idx, value=name)
    row_no = start_row + 1
    for row in rows:
        for c_idx, name in enumerate(columns, start=1):
            ws.cell(row=row_no, column=c_idx, value=row.get(name))
        row_no += 1
    last = row_no - 1
    _finalize_table(ws, start_row, len(columns), last, columns)
    return last


def _raw_sheet_rows(rows: list[ReconRow]) -> tuple[list[str], list[dict[str, Any]]]:
    headers: list[str] = []
    for row in rows[:1]:
        headers = list(row.data.keys())
    return headers, [r.data for r in rows]


def _working_payload(working_rows: list[WorkingRow]) -> list[dict[str, Any]]:
    return [w.values for w in working_rows]


def _is_special(row: ReconRow) -> bool:
    return row.record_kind not in ("Transaction", "Header", "Footer", "")


# --------------------------------------------------------------------------- #
# Summary
# --------------------------------------------------------------------------- #
def _write_summary(
    ws,
    summary: SummaryLayout,
    *,
    identity_values: dict[str, str],
    org_working_sheet: str,
    party_working_sheet: str,
    label_review_counts: dict[str, int],
) -> tuple[list[str], dict[str, int], dict[str, int]]:
    """Render the styled, formula-driven Summary sheet.

    Returns (formula_cells, kind_row_map, label_row_map).
    """
    n_cols = len(summary.columns)

    # Banner.
    ws["A1"] = f"RECONCILIATION SUMMARY — {identity_values.get('pair_id', '')}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    st.style_banner(ws, 1, n_cols)

    # Identity block.
    r = 2
    id_rows: list[int] = []
    for caption, key in summary.identity_fields:
        ws.cell(row=r, column=1, value=caption)
        ws.cell(row=r, column=2, value=identity_values.get(key, ""))
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        ws.cell(row=r, column=2).alignment = st.ALIGN_LEFT
        id_rows.append(r)
        r += 1
    st.style_key_value_block(ws, id_rows)

    header_row = r + 1  # one blank row between identity block and table
    for c_idx, name in enumerate(summary.columns, start=1):
        ws.cell(row=header_row, column=c_idx, value=name)
    st.style_table_header(ws, header_row, n_cols)

    # Column letters for source sheets.
    w_net = _col_letter(WORKING_COLUMNS, "net_org_perspective")
    w_type = _col_letter(WORKING_COLUMNS, "type_label")
    m_org = _col_letter(MATCH_COLUMNS, "org_amount")
    m_party = _col_letter(MATCH_COLUMNS, "party_amount")
    m_status = _col_letter(MATCH_COLUMNS, "match_status")
    m_review = _col_letter(MATCH_COLUMNS, "review_required")

    def sumifs_label(sheet: str, label: str, col: str = w_net) -> str:
        return f"=SUMIFS('{sheet}'!${col}:${col},'{sheet}'!${w_type}:${w_type},\"{label}\")"

    def sumifs_status(col: str, status: str) -> str:
        return (
            f"=SUMIFS({MATCH_EVIDENCE_SHEET}!${col}:${col},"
            f"{MATCH_EVIDENCE_SHEET}!${m_status}:${m_status},\"{status}\")"
        )

    def sumifs_review(col: str) -> str:
        return (
            f"=SUMIFS({MATCH_EVIDENCE_SHEET}!${col}:${col},"
            f"{MATCH_EVIDENCE_SHEET}!${m_review}:${m_review},TRUE)"
        )

    def sumifs_accepted(col: str) -> str:
        return (
            f"={sumifs_status(col, 'matched_strong')[1:]}"
            f"+{sumifs_status(col, 'matched_supported')[1:]}"
            f"+{sumifs_status(col, 'matched_ai')[1:]}"
        )

    kind_row_map: dict[str, int] = {}
    label_row_map: dict[str, int] = {}
    formula_cells: list[str] = []
    data_start = header_row + 1
    row_no = data_start

    # Track range of opening + labels for balance computation.
    first_movement_row = None
    last_movement_row = None

    for srow in summary.rows:
        ws.cell(row=row_no, column=1, value=srow.serial)
        ws.cell(row=row_no, column=2, value=srow.particulars)
        kind_row_map[srow.kind if srow.kind not in ("label",) else f"label::{srow.label}"] = row_no
        if srow.kind == "label":
            label_row_map[srow.label] = row_no

        review_status = srow.review_status
        remarks = srow.remarks

        if srow.kind in {"opening", "label", "closing"}:
            ws[f"{COL_ORG}{row_no}"] = sumifs_label(org_working_sheet, srow.label)
            ws[f"{COL_PARTY}{row_no}"] = sumifs_label(party_working_sheet, srow.label)
            ws[f"{COL_DIFF}{row_no}"] = f"={COL_ORG}{row_no}-{COL_PARTY}{row_no}"
            formula_cells += [f"{COL_ORG}{row_no}", f"{COL_PARTY}{row_no}", f"{COL_DIFF}{row_no}"]
            if srow.kind in {"opening", "label"}:
                if first_movement_row is None:
                    first_movement_row = row_no
                last_movement_row = row_no
            # Annexure hyperlink.
            if srow.annexure:
                cell = ws[f"{COL_ANNEX}{row_no}"]
                cell.value = srow.annexure
                cell.hyperlink = f"#'{srow.annexure}'!A1"
                cell.font = st.Font(name="Calibri", size=10, color="0563C1", underline="single")
            rc = label_review_counts.get(srow.label, 0)
            if rc:
                review_status = "Review required"
                remarks = "See Review_Queue."
        elif srow.kind == "balance_org":
            ws[f"{COL_ORG}{row_no}"] = (
                f"=SUM({COL_ORG}{first_movement_row}:{COL_ORG}{last_movement_row})"
            )
            formula_cells.append(f"{COL_ORG}{row_no}")
        elif srow.kind == "balance_party":
            ws[f"{COL_PARTY}{row_no}"] = (
                f"=SUM({COL_PARTY}{first_movement_row}:{COL_PARTY}{last_movement_row})"
            )
            formula_cells.append(f"{COL_PARTY}{row_no}")
        elif srow.kind == "matched":
            ws[f"{COL_ORG}{row_no}"] = sumifs_accepted(m_org)
            ws[f"{COL_PARTY}{row_no}"] = sumifs_accepted(m_party)
            ws[f"{COL_DIFF}{row_no}"] = f"={COL_ORG}{row_no}-{COL_PARTY}{row_no}"
            formula_cells += [f"{COL_ORG}{row_no}", f"{COL_PARTY}{row_no}", f"{COL_DIFF}{row_no}"]
        elif srow.kind == "candidate_review":
            ws[f"{COL_ORG}{row_no}"] = sumifs_status(m_org, "candidate_review")
            ws[f"{COL_PARTY}{row_no}"] = sumifs_status(m_party, "candidate_review")
            ws[f"{COL_DIFF}{row_no}"] = f"={COL_ORG}{row_no}-{COL_PARTY}{row_no}"
            formula_cells += [f"{COL_ORG}{row_no}", f"{COL_PARTY}{row_no}", f"{COL_DIFF}{row_no}"]
        elif srow.kind == "unmatched_org":
            ws[f"{COL_ORG}{row_no}"] = sumifs_status(m_org, "unmatched_org")
            ws[f"{COL_PARTY}{row_no}"] = sumifs_status(m_party, "unmatched_org")
            ws[f"{COL_DIFF}{row_no}"] = f"={COL_ORG}{row_no}-{COL_PARTY}{row_no}"
            formula_cells += [f"{COL_ORG}{row_no}", f"{COL_PARTY}{row_no}", f"{COL_DIFF}{row_no}"]
        elif srow.kind == "unmatched_party":
            ws[f"{COL_ORG}{row_no}"] = sumifs_status(m_org, "unmatched_party")
            ws[f"{COL_PARTY}{row_no}"] = sumifs_status(m_party, "unmatched_party")
            ws[f"{COL_DIFF}{row_no}"] = f"={COL_ORG}{row_no}-{COL_PARTY}{row_no}"
            formula_cells += [f"{COL_ORG}{row_no}", f"{COL_PARTY}{row_no}", f"{COL_DIFF}{row_no}"]
        elif srow.kind == "review_pending":
            ws[f"{COL_ORG}{row_no}"] = sumifs_review(m_org)
            ws[f"{COL_PARTY}{row_no}"] = sumifs_review(m_party)
            ws[f"{COL_DIFF}{row_no}"] = f"={COL_ORG}{row_no}-{COL_PARTY}{row_no}"
            formula_cells += [f"{COL_ORG}{row_no}", f"{COL_PARTY}{row_no}", f"{COL_DIFF}{row_no}"]
        elif srow.kind == "net_diff":
            borg = kind_row_map.get("balance_org")
            bparty = kind_row_map.get("balance_party")
            ws[f"{COL_ORG}{row_no}"] = f"={COL_ORG}{borg}"
            ws[f"{COL_PARTY}{row_no}"] = f"={COL_PARTY}{bparty}"
            ws[f"{COL_DIFF}{row_no}"] = f"={COL_ORG}{row_no}-{COL_PARTY}{row_no}"
            formula_cells += [f"{COL_ORG}{row_no}", f"{COL_PARTY}{row_no}", f"{COL_DIFF}{row_no}"]
        elif srow.kind == "closing_diff":
            crow = kind_row_map.get("closing")
            ws[f"{COL_ORG}{row_no}"] = f"={COL_ORG}{crow}"
            ws[f"{COL_PARTY}{row_no}"] = f"={COL_PARTY}{crow}"
            ws[f"{COL_DIFF}{row_no}"] = f"={COL_ORG}{row_no}-{COL_PARTY}{row_no}"
            formula_cells += [f"{COL_ORG}{row_no}", f"{COL_PARTY}{row_no}", f"{COL_DIFF}{row_no}"]
        elif srow.kind == "validation_status":
            ws[f"{COL_REVIEW}{row_no}"] = (
                f"=IF(COUNTIF(Validation_Report!${_VALIDATION_STATUS_COL}:"
                f"${_VALIDATION_STATUS_COL},\"FAIL\")=0,\"PASS\",\"FAIL\")"
            )
            formula_cells.append(f"{COL_REVIEW}{row_no}")
        elif srow.kind == "formula_audit_status":
            ws[f"{COL_REVIEW}{row_no}"] = (
                f"=IF(COUNTIF(Formula_Audit!${_FORMULA_AUDIT_STATUS_COL}:"
                f"${_FORMULA_AUDIT_STATUS_COL},\"FAIL\")=0,\"PASS\",\"FAIL\")"
            )
            formula_cells.append(f"{COL_REVIEW}{row_no}")

        # Review status / remarks text columns.
        if srow.kind not in {"validation_status", "formula_audit_status"}:
            ws[f"{COL_REVIEW}{row_no}"] = review_status
        ws[f"{COL_REMARKS}{row_no}"] = remarks
        # Reviewer comment column intentionally blank.

        # Subtle status fill on the review-status cell only.
        rcell = ws[f"{COL_REVIEW}{row_no}"]
        if review_status == "Review required":
            rcell.fill = st.FILL_REVIEW
        elif review_status == "Strong matched":
            rcell.fill = st.FILL_MATCH

        row_no += 1

    last = row_no - 1
    st.apply_borders(ws, header_row, last, n_cols)
    st.apply_number_formats(ws, summary.columns, data_start, last)
    st.freeze(ws, f"A{header_row + 1}")
    st.autosize(ws, n_cols, last)
    # Widen remarks / particulars a little for readability.
    ws.column_dimensions[COL_PARTICULARS].width = max(
        ws.column_dimensions[COL_PARTICULARS].width or 0, 26
    )
    ws.column_dimensions[COL_REMARKS].width = max(ws.column_dimensions[COL_REMARKS].width or 0, 26)
    ws.column_dimensions[COL_REVIEWER].width = max(
        ws.column_dimensions[COL_REVIEWER].width or 0, 22
    )
    return formula_cells, kind_row_map, label_row_map


# --------------------------------------------------------------------------- #
# Executive summary
# --------------------------------------------------------------------------- #
def _write_executive_summary(
    ws,
    *,
    identity_values: dict[str, str],
    ai_status: dict[str, Any] | None,
    stats: dict[str, Any],
    summary_kind_row: dict[str, int],
    review_status_text: str,
) -> None:
    m_status = _col_letter(MATCH_COLUMNS, "match_status")
    span = 4

    ws["A1"] = "EXECUTIVE SUMMARY"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    st.style_banner(ws, 1, span)

    ai_used = "Yes" if (ai_status and ai_status.get("available")) else "No"
    ai_mode = (ai_status or {}).get("mode", "n/a") if ai_status else "n/a"
    review_count_formula = "MAX(0,COUNTA(Review_Queue!$A:$A)-1)"

    def s_ref(kind: str, col: str) -> str:
        row = summary_kind_row.get(kind)
        return f"='Summary'!{col}{row}" if row else "=0"

    r = 3

    def section(title: str) -> None:
        nonlocal r
        ws.cell(row=r, column=1, value=title)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)
        st.style_section_title(ws, r, span=span)
        r += 1

    def kv(caption: str, value: Any, *, is_formula: bool = False, amount: bool = False) -> None:
        nonlocal r
        ws.cell(row=r, column=1, value=caption).font = st.FONT_LABEL
        cell = ws.cell(row=r, column=2, value=value)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=span)
        cell.alignment = st.ALIGN_LEFT
        if amount:
            cell.number_format = st.FMT_AMOUNT
        r += 1

    section("Workbook Identity")
    kv("Pair ID", identity_values.get("pair_id", ""))
    kv("Org ledger name", identity_values.get("org_ledger", ""))
    kv("Party ledger name", identity_values.get("party_name", ""))
    kv("Recon period", identity_values.get("recon_period", ""))
    kv("Generated timestamp", identity_values.get("generated_date", ""))
    kv("Source formalized workbook", identity_values.get("source_formalized", ""))
    kv("Reference workbook used", identity_values.get("reference_workbook", ""))
    kv("AI used", ai_used)
    kv("AI mode", ai_mode)
    kv("Human review required", f"=IF({review_count_formula}>0,\"YES\",\"NO\")", is_formula=True)
    r += 1

    section("Reconciliation Status Snapshot")
    kv("Strong matched count",
       f"=COUNTIF({MATCH_EVIDENCE_SHEET}!${m_status}:${m_status},\"matched_strong\")", is_formula=True)
    kv("Supported matched count",
       f"=COUNTIF({MATCH_EVIDENCE_SHEET}!${m_status}:${m_status},\"matched_supported\")", is_formula=True)
    kv("AI accepted match count",
       f"=COUNTIF({MATCH_EVIDENCE_SHEET}!${m_status}:${m_status},\"matched_ai\")", is_formula=True)
    kv("Candidate review count",
       f"=COUNTIF({MATCH_EVIDENCE_SHEET}!${m_status}:${m_status},\"candidate_review\")", is_formula=True)
    kv("Unmatched org count",
       f"=COUNTIF({MATCH_EVIDENCE_SHEET}!${m_status}:${m_status},\"unmatched_org\")", is_formula=True)
    kv("Unmatched party count",
       f"=COUNTIF({MATCH_EVIDENCE_SHEET}!${m_status}:${m_status},\"unmatched_party\")", is_formula=True)
    kv("Review queue count", f"={review_count_formula}", is_formula=True)
    kv("Validation fail count",
       f"=COUNTIF(Validation_Report!${_VALIDATION_STATUS_COL}:${_VALIDATION_STATUS_COL},\"FAIL\")",
       is_formula=True)
    kv("Formula audit fail count",
       f"=COUNTIF(Formula_Audit!${_FORMULA_AUDIT_STATUS_COL}:${_FORMULA_AUDIT_STATUS_COL},\"FAIL\")",
       is_formula=True)
    kv("Rows dropped", stats.get("rows_dropped", 0))
    kv("Special rows captured", stats.get("special_rows", 0))
    kv("Closing balance difference", s_ref("closing_diff", COL_DIFF), is_formula=True, amount=True)
    r += 1

    section("Amount Snapshot")
    kv("Org ledger total (org perspective)", s_ref("balance_org", COL_ORG), is_formula=True, amount=True)
    kv("Party ledger total (org perspective)", s_ref("balance_party", COL_PARTY), is_formula=True, amount=True)
    kv("Matched amount", s_ref("matched", COL_ORG), is_formula=True, amount=True)
    kv("Candidate review amount", s_ref("candidate_review", COL_ORG), is_formula=True, amount=True)
    u_org = summary_kind_row.get("unmatched_org")
    u_party = summary_kind_row.get("unmatched_party")
    unmatched_formula = (
        f"='Summary'!{COL_ORG}{u_org}+'Summary'!{COL_PARTY}{u_party}"
        if u_org and u_party else "=0"
    )
    kv("Unmatched amount", unmatched_formula, is_formula=True, amount=True)
    kv("Net difference", s_ref("net_diff", COL_DIFF), is_formula=True, amount=True)
    kv("Closing balance check", s_ref("closing", COL_DIFF), is_formula=True, amount=True)
    r += 1

    section("Submission Note")
    ws.cell(row=r, column=1, value=review_status_text).font = st.FONT_LABEL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)
    r += 1
    note = (
        "This workbook is generated from source ledger PDFs with traceable "
        "extracted rows, formula-backed summaries, deterministic-first matching, "
        "bounded AI arbitration, validation checks, and review queues. AI-made "
        "placements remain visible and pass deterministic post-validation. Open in Excel to recalculate "
        "formulas. Final accounting closure requires human approval of all "
        "Review_Queue items."
    )
    cell = ws.cell(row=r, column=1, value=note)
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 4, end_column=span)
    cell.alignment = st.ALIGN_LEFT_WRAP
    cell.font = st.FONT_NOTE

    ws.column_dimensions["A"].width = 34
    for col in ("B", "C", "D"):
        ws.column_dimensions[col].width = 26
    st.freeze(ws, "A3")


# --------------------------------------------------------------------------- #
# Assumptions & limits
# --------------------------------------------------------------------------- #
def _write_assumptions(ws) -> None:
    ws["A1"] = "ASSUMPTIONS & LIMITS"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    st.style_banner(ws, 1, 3)
    items = [
        "Deterministic matching runs first and preserves strong matches.",
        "AI may make final placements only inside bounded reconciliation arbitration.",
        "Every AI placement is restricted to supplied row IDs and deterministically post-validated.",
        "Amounts and references are never invented; absent source values stay absent.",
        "Rejected, invalid, or low-confidence AI decisions remain in Review_Queue.",
        "The Review_Queue contains unresolved work; accepted AI decisions stay visible in audit sheets.",
        "The workbook is formula-backed; open it in Excel to calculate formulas.",
        "Final accounting closure requires human approval of all review-queue items.",
        "Source ledger PDFs remain the source of truth for every figure.",
        "This is a submission-quality, review-ready reconciliation workbook, not a "
        "statement of final audited closure while review items remain.",
    ]
    ws["A3"] = "Caveat"
    st.style_section_title(ws, 3, span=3)
    r = 4
    for idx, text in enumerate(items, start=1):
        ws.cell(row=r, column=1, value=idx).font = st.FONT_LABEL
        cell = ws.cell(row=r, column=2, value=text)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        cell.alignment = st.ALIGN_LEFT_WRAP
        cell.border = st.BORDER_THIN
        ws.cell(row=r, column=1).border = st.BORDER_THIN
        r += 1
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 30
    st.freeze(ws, "A4")


# --------------------------------------------------------------------------- #
# Annexure
# --------------------------------------------------------------------------- #
def _write_annexure(
    wb,
    plan: AnnexurePlan,
    *,
    org_working: list[WorkingRow],
    party_working: list[WorkingRow],
    matches: list[MatchRecord],
    org_working_sheet: str,
    party_working_sheet: str,
) -> list[FormulaAuditItem]:
    audit: list[FormulaAuditItem] = []
    ws = wb.create_sheet(plan.sheet_name)

    w_net = _col_letter(WORKING_COLUMNS, "net_org_perspective")
    w_type = _col_letter(WORKING_COLUMNS, "type_label")
    m_org = _col_letter(MATCH_COLUMNS, "org_amount")
    m_type = _col_letter(MATCH_COLUMNS, "type_label")
    m_status = _col_letter(MATCH_COLUMNS, "match_status")
    m_review = _col_letter(MATCH_COLUMNS, "review_required")

    ws["A1"] = f"Annexure — {plan.label}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    st.style_banner(ws, 1, 6)

    # Section A: label summary block.
    ws["A3"] = "Section A: Label Summary"
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)
    st.style_section_title(ws, 3, span=2)
    metrics = [
        ("Label", plan.label, False),
        ("Org amount",
         f"=SUMIFS('{org_working_sheet}'!${w_net}:${w_net},"
         f"'{org_working_sheet}'!${w_type}:${w_type},\"{plan.label}\")", True),
        ("Party amount",
         f"=SUMIFS('{party_working_sheet}'!${w_net}:${w_net},"
         f"'{party_working_sheet}'!${w_type}:${w_type},\"{plan.label}\")", True),
        ("Difference", "=B5-B6", True),
        ("Accepted matched amount",
         f"=SUMIFS({MATCH_EVIDENCE_SHEET}!${m_org}:${m_org},"
         f"{MATCH_EVIDENCE_SHEET}!${m_type}:${m_type},\"{plan.label}\","
         f"{MATCH_EVIDENCE_SHEET}!${m_status}:${m_status},\"matched_strong\")"
         f"+SUMIFS({MATCH_EVIDENCE_SHEET}!${m_org}:${m_org},"
         f"{MATCH_EVIDENCE_SHEET}!${m_type}:${m_type},\"{plan.label}\","
         f"{MATCH_EVIDENCE_SHEET}!${m_status}:${m_status},\"matched_supported\")"
         f"+SUMIFS({MATCH_EVIDENCE_SHEET}!${m_org}:${m_org},"
         f"{MATCH_EVIDENCE_SHEET}!${m_type}:${m_type},\"{plan.label}\","
         f"{MATCH_EVIDENCE_SHEET}!${m_status}:${m_status},\"matched_ai\")", True),
        ("Candidate review amount",
         f"=SUMIFS({MATCH_EVIDENCE_SHEET}!${m_org}:${m_org},"
         f"{MATCH_EVIDENCE_SHEET}!${m_type}:${m_type},\"{plan.label}\","
         f"{MATCH_EVIDENCE_SHEET}!${m_status}:${m_status},\"candidate_review\")", True),
        ("Unmatched amount",
         f"=SUMIFS({MATCH_EVIDENCE_SHEET}!${m_org}:${m_org},"
         f"{MATCH_EVIDENCE_SHEET}!${m_type}:${m_type},\"{plan.label}\","
         f"{MATCH_EVIDENCE_SHEET}!${m_status}:${m_status},\"unmatched_org\")", True),
        ("Review count",
         f"=COUNTIFS({MATCH_EVIDENCE_SHEET}!${m_type}:${m_type},\"{plan.label}\","
         f"{MATCH_EVIDENCE_SHEET}!${m_review}:${m_review},TRUE)", False),
    ]
    metric_formula_cells: list[str] = []
    r = 4
    for caption, value, amount in metrics:
        ws.cell(row=r, column=1, value=caption).font = st.FONT_LABEL
        cell = ws.cell(row=r, column=2, value=value)
        ws.cell(row=r, column=1).border = st.BORDER_THIN
        cell.border = st.BORDER_THIN
        if amount:
            cell.number_format = st.FMT_AMOUNT
        if isinstance(value, str) and value.startswith("="):
            metric_formula_cells.append(f"B{r}")
        r += 1
    audit.extend(audit_formula_cells(ws, metric_formula_cells, "annexure_summary"))

    # Section B: org-side transactions.
    r += 1
    ws.cell(row=r, column=1, value="Section B: Org-side transactions")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(ANNEX_ORG_HEADERS))
    st.style_section_title(ws, r, span=len(ANNEX_ORG_HEADERS))
    header_b = r + 1
    for c_idx, key in enumerate(ANNEX_ORG_HEADERS, start=1):
        ws.cell(row=header_b, column=c_idx, value=key)
    st.style_table_header(ws, header_b, len(ANNEX_ORG_HEADERS))
    r = header_b + 1
    org_label_rows = [w.values for w in org_working if w.values.get("type_label") == plan.label]
    if not org_label_rows:
        ws.cell(row=r, column=1, value="No transactions found").font = st.FONT_NOTE
        r += 1
    else:
        for row in org_label_rows:
            for c_idx, key in enumerate(ANNEX_ORG_HEADERS, start=1):
                ws.cell(row=r, column=c_idx, value=row.get(key))
            r += 1
        st.apply_borders(ws, header_b, r - 1, len(ANNEX_ORG_HEADERS))
        st.apply_number_formats(ws, ANNEX_ORG_HEADERS, header_b + 1, r - 1)

    # Section C: party-side transactions.
    r += 1
    ws.cell(row=r, column=1, value="Section C: Party-side transactions")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(ANNEX_PARTY_HEADERS))
    st.style_section_title(ws, r, span=len(ANNEX_PARTY_HEADERS))
    header_c = r + 1
    for c_idx, key in enumerate(ANNEX_PARTY_HEADERS, start=1):
        ws.cell(row=header_c, column=c_idx, value=key)
    st.style_table_header(ws, header_c, len(ANNEX_PARTY_HEADERS))
    r = header_c + 1
    party_label_rows = [w.values for w in party_working if w.values.get("type_label") == plan.label]
    if not party_label_rows:
        ws.cell(row=r, column=1, value="No transactions found").font = st.FONT_NOTE
        r += 1
    else:
        for row in party_label_rows:
            for c_idx, key in enumerate(ANNEX_PARTY_HEADERS, start=1):
                ws.cell(row=r, column=c_idx, value=row.get(key))
            r += 1
        st.apply_borders(ws, header_c, r - 1, len(ANNEX_PARTY_HEADERS))
        st.apply_number_formats(ws, ANNEX_PARTY_HEADERS, header_c + 1, r - 1)

    # Section D: match / candidate view.
    r += 1
    ws.cell(row=r, column=1, value="Section D: Match / Candidate view")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(ANNEX_MATCH_HEADERS))
    st.style_section_title(ws, r, span=len(ANNEX_MATCH_HEADERS))
    header_d = r + 1
    for c_idx, key in enumerate(ANNEX_MATCH_HEADERS, start=1):
        ws.cell(row=header_d, column=c_idx, value=key)
    st.style_table_header(ws, header_d, len(ANNEX_MATCH_HEADERS))
    r = header_d + 1
    match_rows = [m.as_dict() for m in matches if m.type_label == plan.label]
    a_org = _col_letter(ANNEX_MATCH_HEADERS, "org_amount")
    a_party = _col_letter(ANNEX_MATCH_HEADERS, "party_amount")
    a_diff = _col_letter(ANNEX_MATCH_HEADERS, "amount_difference")
    diff_cells: list[str] = []
    if not match_rows:
        ws.cell(row=r, column=1, value="No transactions found").font = st.FONT_NOTE
        r += 1
    else:
        for row in match_rows:
            for c_idx, key in enumerate(ANNEX_MATCH_HEADERS, start=1):
                if key in ("reviewer_comment", "manual_status"):
                    ws.cell(row=r, column=c_idx, value="")  # reviewer-owned, blank
                    continue
                if key == "amount_difference":
                    continue  # written as formula below
                ws.cell(row=r, column=c_idx, value=row.get(key))
            ws[f"{a_diff}{r}"] = (
                f"=IF(OR({a_org}{r}=\"\",{a_party}{r}=\"\"),\"\","
                f"ABS(ABS({a_org}{r})-ABS({a_party}{r})))"
            )
            diff_cells.append(f"{a_diff}{r}")
            r += 1
        st.apply_borders(ws, header_d, r - 1, len(ANNEX_MATCH_HEADERS))
        st.apply_number_formats(ws, ANNEX_MATCH_HEADERS, header_d + 1, r - 1)
        audit.extend(
            audit_formula_cells(
                ws, diff_cells, "annexure_match_amount_difference",
                expected_contains=[a_org, a_party],
            )
        )

    st.freeze(ws, "A4")
    st.autosize(ws, max(len(ANNEX_MATCH_HEADERS), 2), r + 1)
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 16)
    return audit


# --------------------------------------------------------------------------- #
# Match evidence and AI decision audit
# --------------------------------------------------------------------------- #
def _write_match_evidence(ws, matches: list[MatchRecord]) -> list[FormulaAuditItem]:
    audit: list[FormulaAuditItem] = []
    match_payload = [m.as_dict() for m in matches]
    # Reviewer-owned columns must be blank regardless of any upstream content.
    for row in match_payload:
        row["reviewer_comment"] = ""
        row["manual_status"] = ""
    match_end = _write_rows(ws, 1, MATCH_COLUMNS, match_payload)

    m_org = _col_letter(MATCH_COLUMNS, "org_amount")
    m_party = _col_letter(MATCH_COLUMNS, "party_amount")
    m_diff = _col_letter(MATCH_COLUMNS, "amount_difference")
    m_status = _col_letter(MATCH_COLUMNS, "match_status")
    m_review = _col_letter(MATCH_COLUMNS, "review_required")

    diff_cells: list[str] = []
    for r in range(2, match_end + 1):
        ws[f"{m_diff}{r}"] = (
            f"=IF(OR({m_org}{r}=\"\",{m_party}{r}=\"\"),\"\","
            f"ABS(ABS({m_org}{r})-ABS({m_party}{r})))"
        )
        diff_cells.append(f"{m_diff}{r}")
    audit.extend(
        audit_formula_cells(
            ws, diff_cells, "match_amount_difference", expected_contains=[m_org, m_party]
        )
    )
    protected_cells: list[str] = []
    for name in PROTECTED_MATCH_COLUMNS:
        letter = _col_letter(MATCH_COLUMNS, name)
        protected_cells.extend(f"{letter}{r}" for r in range(2, match_end + 1))
    if protected_cells:
        audit.extend(audit_protected_cells(ws, protected_cells, "match_protected_evidence"))

    # Subtle status fills (review-safe colour coding).
    status_idx = MATCH_COLUMNS.index("match_status")
    review_idx = MATCH_COLUMNS.index("review_required")
    for r in range(2, match_end + 1):
        status = str(ws.cell(row=r, column=status_idx + 1).value or "")
        review = bool(ws.cell(row=r, column=review_idx + 1).value)
        fill = st.status_fill(status, review)
        if fill is not None:
            for c in range(1, len(MATCH_COLUMNS) + 1):
                ws.cell(row=r, column=c).fill = fill
    return audit


def _write_ai_decision_audit(ws, matches: list[MatchRecord]) -> None:
    """Render authority, validation, and fingerprint metadata for every decision."""
    _write_rows(ws, 1, AI_AUDIT_COLUMNS, [match.as_dict() for match in matches])


# --------------------------------------------------------------------------- #
# Unknown / Needs Review annexure (only written when unknown rows exist)
# --------------------------------------------------------------------------- #
def _write_unknown_needs_review(
    ws,
    groups: list[dict[str, Any]],
    org_unknown: list[dict[str, Any]],
    party_unknown: list[dict[str, Any]],
) -> None:
    """Render the unknown-label grouping report and unclassified transactions.

    Pure presentation of deterministic evidence: a grouping report (Section A)
    plus the unclassified org/party transactions (Sections B/C). No formulas are
    written here, so source evidence is never overwritten.
    """
    ws["A1"] = "UNKNOWN / NEEDS REVIEW — unclassified transaction labels"
    span = max(len(UNKNOWN_GROUPING_COLUMNS), len(UNKNOWN_TX_HEADERS))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    st.style_banner(ws, 1, span)

    ws["A2"] = (
        "These rows did not match a predefined canonical label. They are NEVER "
        "force-matched; each group requires human review."
    )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    ws["A2"].font = st.FONT_NOTE

    # Section A: grouping report.
    r = 4
    ws.cell(row=r, column=1, value="Section A: Unknown-label grouping report")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(UNKNOWN_GROUPING_COLUMNS))
    st.style_section_title(ws, r, span=len(UNKNOWN_GROUPING_COLUMNS))
    header_a = r + 1
    for c_idx, key in enumerate(UNKNOWN_GROUPING_COLUMNS, start=1):
        ws.cell(row=header_a, column=c_idx, value=key)
    st.style_table_header(ws, header_a, len(UNKNOWN_GROUPING_COLUMNS))
    r = header_a + 1
    if not groups:
        ws.cell(row=r, column=1, value="No unknown labels").font = st.FONT_NOTE
        r += 1
    else:
        for row in groups:
            for c_idx, key in enumerate(UNKNOWN_GROUPING_COLUMNS, start=1):
                ws.cell(row=r, column=c_idx, value=row.get(key))
            ws.cell(row=r, column=1).fill = st.FILL_REVIEW
            r += 1
        st.apply_borders(ws, header_a, r - 1, len(UNKNOWN_GROUPING_COLUMNS))

    def _tx_section(title: str, rows: list[dict[str, Any]], start: int) -> int:
        ws.cell(row=start, column=1, value=title)
        ws.merge_cells(start_row=start, start_column=1, end_row=start,
                       end_column=len(UNKNOWN_TX_HEADERS))
        st.style_section_title(ws, start, span=len(UNKNOWN_TX_HEADERS))
        header = start + 1
        for c_idx, key in enumerate(UNKNOWN_TX_HEADERS, start=1):
            ws.cell(row=header, column=c_idx, value=key)
        st.style_table_header(ws, header, len(UNKNOWN_TX_HEADERS))
        rr = header + 1
        if not rows:
            ws.cell(row=rr, column=1, value="No transactions found").font = st.FONT_NOTE
            return rr + 1
        for row in rows:
            for c_idx, key in enumerate(UNKNOWN_TX_HEADERS, start=1):
                ws.cell(row=rr, column=c_idx, value=row.get(key))
            rr += 1
        st.apply_borders(ws, header, rr - 1, len(UNKNOWN_TX_HEADERS))
        st.apply_number_formats(ws, UNKNOWN_TX_HEADERS, header + 1, rr - 1)
        return rr

    r += 1
    r = _tx_section("Section B: Org-side unclassified transactions", org_unknown, r)
    r += 1
    r = _tx_section("Section C: Party-side unclassified transactions", party_unknown, r)

    st.freeze(ws, "A5")
    st.autosize(ws, span, r + 1)


# --------------------------------------------------------------------------- #
# Orchestration
# --------------------------------------------------------------------------- #
def _reorder_sheets(wb: Workbook, desired: list[str]) -> None:
    present = [name for name in desired if name in wb.sheetnames]
    remaining = [name for name in wb.sheetnames if name not in present]
    wb._sheets = [wb[name] for name in present + remaining]


def write_reconciliation_workbook(
    *,
    output_path: Path,
    pair_id: str,
    formalized_path: Path,
    reference_path: str,
    labels: list[str],
    org_sheet_name: str,
    party_sheet_name: str,
    org_rows: list[ReconRow],
    party_rows: list[ReconRow],
    org_working: list[WorkingRow],
    party_working: list[WorkingRow],
    summary_layout: SummaryLayout,
    annex_plans: list[AnnexurePlan],
    matches: list[MatchRecord],
    validation_items: list[ValidationItem],
    ai_status: dict[str, Any] | None = None,
    include_master_match_table: bool = False,
    recon_period: str = "",
    org_ledger_name: str | None = None,
    party_ledger_name: str | None = None,
) -> list[FormulaAuditItem]:
    """Write the full reconciliation workbook and return formula audit entries."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    formula_audit: list[FormulaAuditItem] = []
    generated = datetime.now(timezone.utc).isoformat(timespec="seconds")

    org_ledger_name = org_ledger_name or org_sheet_name
    party_ledger_name = party_ledger_name or party_sheet_name

    # Derived stats used by the executive summary.
    special_rows = sum(1 for r in org_rows + party_rows if _is_special(r))
    review_records = [m for m in matches if m.review_required]
    review_queue_rows = build_review_queue(org_rows, party_rows, matches)
    label_review_counts: dict[str, int] = {}
    for m in review_records:
        label_review_counts[m.type_label] = label_review_counts.get(m.type_label, 0) + 1
    review_burden = len(review_queue_rows)
    review_status_text = (
        "Status: review-ready reconciliation — human review pending."
        if review_burden
        else "Status: no outstanding review items — pending human sign-off."
    )

    identity_values = {
        "pair_id": pair_id,
        "party_name": party_ledger_name,
        "org_ledger": org_ledger_name,
        "party_ledger": party_ledger_name,
        "recon_period": recon_period or "Not specified",
        "generated_date": generated,
        "source_formalized": str(formalized_path),
        "reference_workbook": reference_path or "Not found (profile used)",
        "status": review_status_text.replace("Status: ", ""),
    }

    # --- README ---
    ws_readme = wb.create_sheet("README")
    ws_readme["A1"] = "RECONCILIATION WORKBOOK — README"
    ws_readme.merge_cells("A1:B1")
    st.style_banner(ws_readme, 1, 2)
    readme_rows = [
        ("pair_id", pair_id),
        ("generated_utc", generated),
        ("recon_period", identity_values["recon_period"]),
        ("source_formalized_workbook", str(formalized_path)),
        ("reference_workbook", identity_values["reference_workbook"]),
        ("org_ledger_sheet", org_sheet_name),
        ("party_ledger_sheet", party_sheet_name),
        ("labels_used", ", ".join(labels)),
        ("ai_used", "yes" if (ai_status and ai_status.get("available")) else "no"),
        ("ai_mode", (ai_status or {}).get("mode", "n/a")),
        ("ai_detail", (ai_status or {}).get("detail", "AI not invoked.")),
        ("matching", "deterministic-first; bounded AI arbitration after deterministic failure"),
        ("review_queue_rows", review_burden),
        ("output_path", str(output_path)),
        ("caveat", "Manual review fields intentionally blank for human decisions."),
    ]
    for idx, (k, v) in enumerate(readme_rows, start=3):
        ws_readme.cell(row=idx, column=1, value=k).font = st.FONT_LABEL
        ws_readme.cell(row=idx, column=2, value=v)
    st.autosize(ws_readme, 2, len(readme_rows) + 3)
    ws_readme.column_dimensions["B"].width = 60

    # --- Raw + working ledgers ---
    org_headers, org_payload = _raw_sheet_rows(org_rows)
    ws_org = wb.create_sheet(org_sheet_name[:31])
    _write_rows(ws_org, 1, org_headers, org_payload)

    w_net_src = _col_letter(WORKING_COLUMNS, "net_source")
    w_deb_src = _col_letter(WORKING_COLUMNS, "debit_source")
    w_cred_src = _col_letter(WORKING_COLUMNS, "credit_source")
    w_net_org = _col_letter(WORKING_COLUMNS, "net_org_perspective")
    w_deb_org = _col_letter(WORKING_COLUMNS, "debit_org_perspective")
    w_cred_org = _col_letter(WORKING_COLUMNS, "credit_org_perspective")

    def _write_working_formulas(ws, last_row: int) -> None:
        for r in range(2, last_row + 1):
            ws[f"{w_net_src}{r}"] = f"=N({w_deb_src}{r})-N({w_cred_src}{r})"
            ws[f"{w_net_org}{r}"] = f"=N({w_deb_org}{r})-N({w_cred_org}{r})"
        formula_audit.extend(
            audit_formula_cells(
                ws, [f"{w_net_src}{r}" for r in range(2, last_row + 1)],
                "working_net_source", expected_contains=[w_deb_src, w_cred_src],
            )
        )
        formula_audit.extend(
            audit_formula_cells(
                ws, [f"{w_net_org}{r}" for r in range(2, last_row + 1)],
                "working_net_org", expected_contains=[w_deb_org, w_cred_org],
            )
        )

    ws_w_org = wb.create_sheet(f"Working {org_sheet_name}"[:31])
    org_end = _write_rows(ws_w_org, 1, WORKING_COLUMNS, _working_payload(org_working))
    _write_working_formulas(ws_w_org, org_end)

    party_headers, party_payload = _raw_sheet_rows(party_rows)
    ws_party = wb.create_sheet(party_sheet_name[:31])
    _write_rows(ws_party, 1, party_headers, party_payload)

    ws_w_party = wb.create_sheet(f"Working {party_sheet_name}"[:31])
    party_end = _write_rows(ws_w_party, 1, WORKING_COLUMNS, _working_payload(party_working))
    _write_working_formulas(ws_w_party, party_end)

    # --- Summary ---
    ws_summary = wb.create_sheet("Summary")
    summary_formula_cells, summary_kind_row, _label_row_map = _write_summary(
        ws_summary,
        summary_layout,
        identity_values=identity_values,
        org_working_sheet=ws_w_org.title,
        party_working_sheet=ws_w_party.title,
        label_review_counts=label_review_counts,
    )
    formula_audit.extend(audit_formula_cells(ws_summary, summary_formula_cells, "summary_totals"))

    # --- Primary match evidence + AI decision audit ---
    ws_match = wb.create_sheet(MATCH_EVIDENCE_SHEET)
    formula_audit.extend(_write_match_evidence(ws_match, matches))
    ws_ai_audit = wb.create_sheet(AI_DECISION_AUDIT_SHEET)
    _write_ai_decision_audit(ws_ai_audit, matches)
    if include_master_match_table:
        # Optional legacy compatibility sheet. Match_Evidence remains canonical.
        ws_legacy = wb.create_sheet("Master_Match_Table")
        _write_match_evidence(ws_legacy, matches)

    # --- Annexures ---
    for plan in annex_plans:
        formula_audit.extend(
            _write_annexure(
                wb, plan,
                org_working=org_working, party_working=party_working, matches=matches,
                org_working_sheet=ws_w_org.title, party_working_sheet=ws_w_party.title,
            )
        )

    # --- Unknown / Needs Review annexure (only when unclassified rows exist) ---
    unknown_present = has_unknown_rows(org_rows, party_rows)
    if unknown_present:
        ws_unknown = wb.create_sheet(UNKNOWN_REVIEW_SHEET)
        _write_unknown_needs_review(
            ws_unknown,
            build_unknown_label_groups(org_rows, party_rows),
            unknown_rows_for_side(org_rows),
            unknown_rows_for_side(party_rows),
        )

    # --- Review queue ---
    ws_review = wb.create_sheet("Review_Queue")
    _write_rows(ws_review, 1, REVIEW_QUEUE_COLUMNS, review_queue_rows)
    # Light-yellow band over the priority column to signal action area.
    if review_queue_rows:
        for r in range(2, len(review_queue_rows) + 2):
            ws_review.cell(row=r, column=1).fill = st.FILL_REVIEW

    # --- Validation ---
    ws_val = wb.create_sheet("Validation_Report")
    _write_rows(
        ws_val, 1, ["check", "scope", "status", "count", "details"],
        [v.as_dict() for v in validation_items],
    )

    # --- Formula audit ---
    ws_fa = wb.create_sheet("Formula_Audit")
    fa_headers = [
        "sheet_name", "cell", "expected_formula_category",
        "formula_present", "formula_text", "status", "issue",
    ]
    _write_rows(ws_fa, 1, fa_headers, [f.as_dict() for f in formula_audit])

    # --- Assumptions & limits ---
    ws_assume = wb.create_sheet("Assumptions_And_Limits")
    _write_assumptions(ws_assume)

    # --- Executive summary (built last; references computed Summary rows) ---
    ws_exec = wb.create_sheet("Executive_Summary")
    _write_executive_summary(
        ws_exec,
        identity_values=identity_values,
        ai_status=ai_status,
        stats={"rows_dropped": 0, "special_rows": special_rows},
        summary_kind_row=summary_kind_row,
        review_status_text=review_status_text,
    )

    # --- Final sheet ordering ---
    desired = (
        REQUIRED_SHEET_ORDER_HEAD
        + [org_sheet_name[:31], f"Working {org_sheet_name}"[:31],
           party_sheet_name[:31], f"Working {party_sheet_name}"[:31],
           "Summary", AI_DECISION_AUDIT_SHEET, MATCH_EVIDENCE_SHEET]
        + (["Master_Match_Table"] if include_master_match_table else [])
        + [p.sheet_name for p in annex_plans]
        + ([UNKNOWN_REVIEW_SHEET] if unknown_present else [])
        + ["Review_Queue", "Validation_Report", "Formula_Audit", "Assumptions_And_Limits"]
    )
    _reorder_sheets(wb, desired)

    # Print/page setup modeled on the reference workbook (landscape, fit-to-width).
    for ws in wb.worksheets:
        st.apply_page_setup(ws)

    wb.save(output_path)
    return formula_audit
