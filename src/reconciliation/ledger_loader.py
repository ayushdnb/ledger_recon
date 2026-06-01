"""Load formalized workbook ledgers for reconciliation."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.reconciliation.recon_models import ReconRow

SPECIAL_SHEETS = {
    "README",
    "Extraction_Audit",
    "Page_Analysis",
    "Validation_Report",
    "Review_Queue",
}


@dataclass
class LoadedLedgers:
    """Both formalized ledgers and workbook metadata."""

    workbook_path: Path
    sheet_names: list[str]
    org_sheet: str
    party_sheet: str
    org_rows: list[ReconRow]
    party_rows: list[ReconRow]


def _rows_from_sheet(workbook_path: Path, sheet_name: str) -> list[ReconRow]:
    wb = load_workbook(workbook_path, data_only=False, read_only=True)
    ws = wb[sheet_name]
    headers = [str(c.value) if c.value is not None else "" for c in ws[1]]
    rows: list[ReconRow] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row_data: dict[str, Any] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            row_data[header] = values[idx] if idx < len(values) else None
        rows.append(ReconRow(data=row_data, sheet_name=sheet_name))
    wb.close()
    return rows


def _infer_role(rows: list[ReconRow]) -> str:
    for row in rows[:20]:
        role = str(row.data.get("ledger_role", "")).strip()
        if role:
            return role
    return ""


def load_formalized_ledgers(workbook_path: Path) -> LoadedLedgers:
    """Load formalized workbook and map main sheets to org/party roles."""
    if not workbook_path.exists():
        raise FileNotFoundError(f"Formalized workbook not found: {workbook_path}")
    wb = load_workbook(workbook_path, data_only=False, read_only=True)
    sheet_names = wb.sheetnames
    main_sheets = [s for s in sheet_names if s not in SPECIAL_SHEETS]
    wb.close()
    if len(main_sheets) != 2:
        raise RuntimeError(
            f"Expected exactly 2 ledger sheets in formalized workbook; found {len(main_sheets)}: {main_sheets}"
        )

    rows_a = _rows_from_sheet(workbook_path, main_sheets[0])
    rows_b = _rows_from_sheet(workbook_path, main_sheets[1])
    role_a = _infer_role(rows_a)
    role_b = _infer_role(rows_b)

    if role_a == "org_ledger" and role_b == "party_ledger":
        org_sheet, party_sheet, org_rows, party_rows = (
            main_sheets[0],
            main_sheets[1],
            rows_a,
            rows_b,
        )
    elif role_b == "org_ledger" and role_a == "party_ledger":
        org_sheet, party_sheet, org_rows, party_rows = (
            main_sheets[1],
            main_sheets[0],
            rows_b,
            rows_a,
        )
    else:
        # Fallback deterministic order if role tags are missing.
        org_sheet, party_sheet, org_rows, party_rows = (
            main_sheets[0],
            main_sheets[1],
            rows_a,
            rows_b,
        )

    return LoadedLedgers(
        workbook_path=workbook_path,
        sheet_names=sheet_names,
        org_sheet=org_sheet,
        party_sheet=party_sheet,
        org_rows=org_rows,
        party_rows=party_rows,
    )

