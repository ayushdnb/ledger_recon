"""Profile an old reference reconciliation workbook (structure only).

Read-only inspection of a reference ``.xlsx`` so a new standardized workbook can
later be made structurally similar. Captures sheet names, used ranges, merged
cells, detected headers, formula counts, non-empty counts, ledger/summary sheet
guesses, and sample rows. Writes a JSON profile and an Excel profile report.

Never mutates the reference workbook.

Run from the project root::

    python -m src.excel.profile_reference_workbook --workbook "data/03_reference_workbooks/old_recon_workbooks/<name>.xlsx"
"""

from __future__ import annotations

import argparse
import json
import logging
import re
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.config import settings
from src.excel.workbook_io import write_workbook

logger = logging.getLogger("profile_reference_workbook")

HEADER_SCAN_ROWS = 20
SAMPLE_ROWS = 20

LEDGER_HEADER_KEYWORDS = (
    "date",
    "debit",
    "credit",
    "particular",
    "voucher",
    "vch",
    "amount",
    "balance",
    "narration",
    "bill",
    "invoice",
    "ledger",
)
SUMMARY_KEYWORDS = (
    "summary",
    "recon",
    "total",
    "abstract",
    "statement",
    "annexure",
    "overview",
)


def slugify(text: str) -> str:
    """Build a filesystem-safe slug from a workbook name.

    Uses the portion before the first ' - ' (a common "Name - Period" pattern)
    when present, then lowercases and replaces non-alphanumerics with single
    underscores. This is a general rule, not a file-specific hack.
    """
    base = text.split(" - ", 1)[0]
    if base == text:
        base = text.split("-", 1)[0]
    slug = re.sub(r"[^a-zA-Z0-9]+", "_", base).strip("_").lower()
    return slug or "reference_workbook"


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def _argb(color: object) -> str:
    """Return an ARGB hex string for an openpyxl color, or '' if absent."""
    rgb = getattr(color, "rgb", None)
    if isinstance(rgb, str) and rgb and rgb != "00000000":
        return rgb
    return ""


def _cell_style(cell) -> dict:
    """Capture the visual style of a single cell (read-only inspection)."""
    font = cell.font
    fill = cell.fill
    fg = _argb(getattr(fill, "fgColor", None)) if fill and fill.fill_type == "solid" else ""
    return {
        "font_name": font.name,
        "font_size": float(font.size) if font.size is not None else None,
        "bold": bool(font.bold),
        "italic": bool(font.italic),
        "font_color": _argb(getattr(font, "color", None)),
        "fill_color": fg,
        "number_format": cell.number_format,
        "horizontal": getattr(cell.alignment, "horizontal", None),
        "wrap_text": bool(getattr(cell.alignment, "wrap_text", False)),
    }


def _detect_header(rows: list[list[object]]) -> tuple[int | None, list[str]]:
    """Pick the most header-like row within the first scanned rows."""
    best_index: int | None = None
    best_count = 0
    for index, row in enumerate(rows):
        non_empty = [c for c in row if _cell_str(c).strip()]
        if len(non_empty) > best_count:
            best_count = len(non_empty)
            best_index = index
    if best_index is None:
        return None, []
    headers = [_cell_str(c).strip() for c in rows[best_index]]
    # Trim trailing empties.
    while headers and not headers[-1]:
        headers.pop()
    return best_index, headers


def _profile_sheet(worksheet) -> dict:
    max_row = worksheet.max_row or 0
    max_col = worksheet.max_column or 0

    scan_rows: list[list[object]] = []
    for row in worksheet.iter_rows(
        min_row=1, max_row=min(HEADER_SCAN_ROWS, max_row), values_only=True
    ):
        scan_rows.append(list(row))

    header_index, headers = _detect_header(scan_rows)

    formula_count = 0
    non_empty_rows = 0
    non_empty_cols: set[int] = set()
    sample_rows: list[list[str]] = []

    for row in worksheet.iter_rows(min_row=1, max_row=max_row):
        row_has_value = False
        for cell in row:
            value = cell.value
            if value is None:
                continue
            if isinstance(value, str) and value.startswith("="):
                formula_count += 1
            text = _cell_str(value).strip()
            if text:
                row_has_value = True
                non_empty_cols.add(cell.column)
        if row_has_value:
            non_empty_rows += 1
            if len(sample_rows) < SAMPLE_ROWS:
                sample_rows.append([_cell_str(c.value) for c in row])

    headers_lower = " ".join(headers).lower()
    name_lower = worksheet.title.lower()
    is_ledger_like = any(k in headers_lower for k in LEDGER_HEADER_KEYWORDS)
    is_summary_like = any(
        k in headers_lower or k in name_lower for k in SUMMARY_KEYWORDS
    )

    # --- Formatting capture (read-only) -----------------------------------
    # Column widths and a bounded window of row heights.
    column_widths = {
        letter: dim.width
        for letter, dim in worksheet.column_dimensions.items()
        if dim.width is not None
    }
    row_heights = {
        str(r): worksheet.row_dimensions[r].height
        for r in range(1, min(max_row, 12) + 1)
        if r in worksheet.row_dimensions and worksheet.row_dimensions[r].height is not None
    }

    setup = worksheet.page_setup
    page_setup = {
        "orientation": getattr(setup, "orientation", None),
        "paper_size": getattr(setup, "paperSize", None),
        "fit_to_width": getattr(setup, "fitToWidth", None),
        "fit_to_height": getattr(setup, "fitToHeight", None),
        "fit_to_page": bool(
            getattr(getattr(worksheet, "sheet_properties", None), "pageSetUpPr", None)
            and worksheet.sheet_properties.pageSetUpPr.fitToPage
        ),
    }

    # Title (A1) and detected header-row cell styles (a bounded sample).
    title_style = _cell_style(worksheet["A1"]) if max_row >= 1 and max_col >= 1 else {}
    header_cell_styles: list[dict] = []
    if header_index is not None:
        header_row_no = header_index + 1
        for col in range(1, min(max_col, 12) + 1):
            cell = worksheet.cell(row=header_row_no, column=col)
            if cell.value is None and not _argb(getattr(cell.fill, "fgColor", None)):
                continue
            header_cell_styles.append(_cell_style(cell))

    return {
        "sheet_name": worksheet.title,
        "used_range": worksheet.dimensions,
        "max_row": max_row,
        "max_column": max_col,
        "merged_cells": [str(r) for r in worksheet.merged_cells.ranges],
        "merged_cell_count": len(worksheet.merged_cells.ranges),
        "header_row_index": (header_index + 1) if header_index is not None else None,
        "headers": headers,
        "formula_count": formula_count,
        "non_empty_row_count": non_empty_rows,
        "non_empty_column_count": len(non_empty_cols),
        "is_ledger_like": is_ledger_like,
        "is_summary_like": is_summary_like,
        "freeze_panes": str(worksheet.freeze_panes or ""),
        "column_widths": column_widths,
        "row_heights": row_heights,
        "page_setup": page_setup,
        "title_style": title_style,
        "header_cell_styles": header_cell_styles,
        "sample_rows": sample_rows,
    }


def profile_workbook(workbook_path: Path) -> dict:
    """Build the structural profile of a workbook (read-only)."""
    if not workbook_path.exists():
        raise FileNotFoundError(f"Reference workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    try:
        sheets = [_profile_sheet(workbook[name]) for name in workbook.sheetnames]
    finally:
        workbook.close()

    return {
        "workbook_name": workbook_path.name,
        "workbook_path": str(workbook_path),
        "generated_utc": datetime.now(timezone.utc).isoformat(),
        "sheet_count": len(sheets),
        "sheet_names": [s["sheet_name"] for s in sheets],
        "ledger_like_sheets": [s["sheet_name"] for s in sheets if s["is_ledger_like"]],
        "summary_like_sheets": [s["sheet_name"] for s in sheets if s["is_summary_like"]],
        "sheets": sheets,
    }


def _build_report_sheets(profile: dict) -> list[tuple[str, pd.DataFrame]]:
    overview_rows = [
        {
            "sheet_name": s["sheet_name"],
            "used_range": s["used_range"],
            "max_row": s["max_row"],
            "max_column": s["max_column"],
            "merged_cell_count": s["merged_cell_count"],
            "header_row_index": s["header_row_index"],
            "formula_count": s["formula_count"],
            "non_empty_row_count": s["non_empty_row_count"],
            "non_empty_column_count": s["non_empty_column_count"],
            "is_ledger_like": s["is_ledger_like"],
            "is_summary_like": s["is_summary_like"],
        }
        for s in profile["sheets"]
    ]
    overview_df = pd.DataFrame(
        overview_rows,
        columns=[
            "sheet_name",
            "used_range",
            "max_row",
            "max_column",
            "merged_cell_count",
            "header_row_index",
            "formula_count",
            "non_empty_row_count",
            "non_empty_column_count",
            "is_ledger_like",
            "is_summary_like",
        ],
    )

    header_rows = []
    for s in profile["sheets"]:
        for position, header in enumerate(s["headers"], start=1):
            header_rows.append(
                {
                    "sheet_name": s["sheet_name"],
                    "header_position": position,
                    "header_text": header,
                }
            )
    headers_df = pd.DataFrame(
        header_rows, columns=["sheet_name", "header_position", "header_text"]
    )

    sample_rows = []
    for s in profile["sheets"]:
        for row_index, row in enumerate(s["sample_rows"], start=1):
            record: dict = {"sheet_name": s["sheet_name"], "sample_row_index": row_index}
            for col_position, value in enumerate(row[:15], start=1):
                record[f"col_{col_position:02d}"] = value
            sample_rows.append(record)
    sample_columns = ["sheet_name", "sample_row_index"] + [
        f"col_{i:02d}" for i in range(1, 16)
    ]
    sample_df = pd.DataFrame(sample_rows, columns=sample_columns)

    return [
        ("Sheets_Overview", overview_df),
        ("Detected_Headers", headers_df),
        ("Sample_Rows", sample_df),
    ]


def run_profile(workbook_path: Path) -> dict[str, Path]:
    """Profile the workbook and write JSON + Excel outputs."""
    profile = profile_workbook(workbook_path)
    slug = slugify(workbook_path.stem)

    out_dir = settings.resolved(settings.reference_profile_output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    json_path = out_dir / f"reference_workbook_profile__{slug}.json"
    xlsx_path = out_dir / f"reference_workbook_profile__{slug}.xlsx"

    json_path.write_text(
        json.dumps(profile, indent=2, ensure_ascii=False), encoding="utf-8"
    )

    readme_text = (
        "REFERENCE WORKBOOK STRUCTURE PROFILE (read-only).\n\n"
        "Structural inspection only: sheet names, used ranges, merged cells, "
        "detected headers, formula counts, non-empty counts, ledger/summary sheet "
        "guesses, and sample rows. No transaction values were interpreted and the "
        "reference workbook was not modified.\n\n"
        f"workbook: {profile['workbook_name']}\n"
        f"generated (UTC): {profile['generated_utc']}\n"
        f"sheet_count: {profile['sheet_count']}\n"
    )
    write_workbook(xlsx_path, readme_text, _build_report_sheets(profile))

    return {"json": json_path, "xlsx": xlsx_path}


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Profile an old reference reconciliation workbook (structure only).",
    )
    parser.add_argument(
        "--workbook",
        required=True,
        help="Path to the reference .xlsx workbook to profile.",
    )
    args = parser.parse_args()

    logging.basicConfig(
        level=getattr(logging, settings.log_level.upper(), logging.INFO),
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    )

    workbook_path = Path(args.workbook)
    if not workbook_path.is_absolute():
        workbook_path = settings.project_root / workbook_path

    outputs = run_profile(workbook_path)
    logger.info("Reference profile JSON: %s", outputs["json"])
    logger.info("Reference profile XLSX: %s", outputs["xlsx"])
    print(f"Reference profile JSON: {outputs['json']}")
    print(f"Reference profile XLSX: {outputs['xlsx']}")


if __name__ == "__main__":
    main()
