"""Read-only inspection of generated reconciliation workbooks.

Audit utility only: opens workbooks with openpyxl in read-only intent and prints
a compact integrity report. It never mutates any workbook.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

REQUIRED_SHEETS = {
    "README",
    "Executive_Summary",
    "Summary",
    "AI_Decision_Audit",
    "Match_Evidence",
    "Review_Queue",
    "Validation_Report",
    "Formula_Audit",
    "Assumptions_And_Limits",
}

REVIEW_MANUAL_COLUMNS = ["reviewer_comment", "manual_status", "resolved_by", "resolved_date"]
REVIEW_REQUIRED_COLUMNS = [
    "priority",
    "source_area",
    "source_sheet",
    "source_row_id",
    "source_file",
    "page_number",
    "source_row_number",
    "type_label",
    "date",
    "reference",
    "normalized_reference",
    "amount",
    "match_status",
    "primary_issue_code",
    "reason",
    "suggested_action",
    "source_raw_text",
]

EVIDENCE_COLUMNS = {
    "org_raw_type",
    "party_raw_type",
    "org_particulars",
    "party_particulars",
    "org_reference",
    "party_reference",
    "org_normalized_reference",
    "party_normalized_reference",
    "raw_text",
    "source_file",
    "source_row_id",
}

MISSING_REF_TOKENS = {"", "none", "nan", "null", "-", "--"}


def _is_formula(value: object) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _is_missing(value: object) -> bool:
    return str(value if value is not None else "").strip().lower() in MISSING_REF_TOKENS


def _col_index(headers: list[str], name: str) -> int | None:
    for idx, header in enumerate(headers, start=1):
        if header == name:
            return idx
    return None


def _configured_label_count() -> int | None:
    """Best-effort read of configured label count for annexure verification."""
    here = Path(__file__).resolve()
    for parent in here.parents:
        candidate = parent / "config" / "type_labels.json"
        if candidate.exists():
            try:
                return len(json.loads(candidate.read_text(encoding="utf-8")))
            except (json.JSONDecodeError, ValueError):
                return None
    return None


def inspect(path: Path) -> dict[str, object]:
    wb = load_workbook(path, data_only=False)
    report: dict[str, object] = {"workbook": str(path), "sheets": wb.sheetnames}

    # Required sheet presence.
    missing = sorted(REQUIRED_SHEETS - set(wb.sheetnames))
    report["required_sheets_present"] = not missing
    report["missing_required_sheets"] = missing
    report["executive_summary_present"] = "Executive_Summary" in wb.sheetnames
    report["assumptions_present"] = "Assumptions_And_Limits" in wb.sheetnames

    # Executive_Summary required sections.
    if "Executive_Summary" in wb.sheetnames:
        ws = wb["Executive_Summary"]
        text = {str(c.value) for row in ws.iter_rows() for c in row if c.value is not None}
        needed = {"Workbook Identity", "Reconciliation Status Snapshot", "Amount Snapshot",
                  "Submission Note"}
        report["executive_summary_sections_ok"] = needed.issubset(text)
        report["executive_summary_formula_count"] = sum(
            1 for row in ws.iter_rows() for c in row if _is_formula(c.value)
        )

    # Summary formulas + top block.
    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]
        formula_cnt = sum(1 for row in ws.iter_rows() for c in row if _is_formula(c.value))
        report["summary_formula_count"] = formula_cnt
        text = {str(c.value) for row in ws.iter_rows() for c in row if c.value is not None}
        report["summary_top_block_ok"] = (
            "Vendor / Party Name" in text and "Recon Period" in text and "Status" in text
        )

    report["annexure_count"] = sum(1 for s in wb.sheetnames if s.startswith("Annex_"))
    expected_labels = _configured_label_count()
    report["configured_label_count"] = expected_labels
    report["annexure_count_matches_labels"] = (
        expected_labels is None or report["annexure_count"] == expected_labels
    )

    # Empty annexures still complete (have all four sections).
    annex_section_ok = True
    for s in wb.sheetnames:
        if not s.startswith("Annex_"):
            continue
        ws = wb[s]
        text = {str(c.value) for row in ws.iter_rows() for c in row if c.value is not None}
        if not (
            any("Section A" in t for t in text)
            and any("Section B" in t for t in text)
            and any("Section C" in t for t in text)
            and any("Section D" in t for t in text)
        ):
            annex_section_ok = False
            break
    report["annexures_complete"] = annex_section_ok

    # Match_Evidence inspection
    evidence = wb["Match_Evidence"]
    headers = [str(c.value) if c.value is not None else "" for c in evidence[1]]
    report["match_evidence_headers"] = headers
    ad_idx = _col_index(headers, "amount_difference")
    report["amount_difference_col"] = get_column_letter(ad_idx) if ad_idx else None

    data_rows = list(evidence.iter_rows(min_row=2))
    report["match_evidence_data_rows"] = len(data_rows)

    # amount_difference formulas where both amounts present
    org_idx = _col_index(headers, "org_amount")
    party_idx = _col_index(headers, "party_amount")
    ad_formula_ok = 0
    ad_formula_missing = 0
    for row in data_rows:
        org_v = row[org_idx - 1].value if org_idx else None
        party_v = row[party_idx - 1].value if party_idx else None
        ad_v = row[ad_idx - 1].value if ad_idx else None
        if org_v is not None and party_v is not None:
            if _is_formula(ad_v):
                ad_formula_ok += 1
            else:
                ad_formula_missing += 1
    report["amount_difference_formula_ok"] = ad_formula_ok
    report["amount_difference_formula_missing"] = ad_formula_missing

    # Evidence columns must not be formulas
    evidence_formula_violations: dict[str, int] = {}
    for name in EVIDENCE_COLUMNS:
        idx = _col_index(headers, name)
        if idx is None:
            continue
        viol = sum(1 for row in data_rows if _is_formula(row[idx - 1].value))
        if viol:
            evidence_formula_violations[name] = viol
    report["evidence_formula_violations"] = evidence_formula_violations

    # Strong matches with missing reference
    status_idx = _col_index(headers, "match_status")
    onr_idx = _col_index(headers, "org_normalized_reference")
    pnr_idx = _col_index(headers, "party_normalized_reference")
    type_idx = _col_index(headers, "type_label")
    strong_missing_ref = 0
    matched_strong = 0
    matched_supported = 0
    candidate_review = 0
    unmatched = 0
    for row in data_rows:
        status = str(row[status_idx - 1].value or "") if status_idx else ""
        if status == "matched_strong":
            matched_strong += 1
            onr = row[onr_idx - 1].value if onr_idx else None
            pnr = row[pnr_idx - 1].value if pnr_idx else None
            if _is_missing(onr) or _is_missing(pnr):
                strong_missing_ref += 1
        elif status == "matched_supported":
            matched_supported += 1
        elif status == "candidate_review":
            candidate_review += 1
        elif status in {"unmatched_org", "unmatched_party"}:
            unmatched += 1
    report["matched_strong"] = matched_strong
    report["matched_supported"] = matched_supported
    report["candidate_review"] = candidate_review
    report["unmatched"] = unmatched
    report["strong_matches_missing_reference"] = strong_missing_ref

    # Transaction rows with type_label Unknown across working ledgers
    unknown_tx = 0
    special_rows = 0
    for sheet in wb.sheetnames:
        if not sheet.startswith("Working "):
            continue
        ws = wb[sheet]
        wheaders = [str(c.value) if c.value is not None else "" for c in ws[1]]
        rk_idx = _col_index(wheaders, "record_kind")
        tl_idx = _col_index(wheaders, "type_label")
        for row in ws.iter_rows(min_row=2):
            rk = str(row[rk_idx - 1].value or "") if rk_idx else ""
            tl = str(row[tl_idx - 1].value or "") if tl_idx else ""
            if rk and rk != "Transaction":
                special_rows += 1
            if rk == "Transaction" and tl == "Unknown":
                unknown_tx += 1
    report["transaction_rows_type_unknown"] = unknown_tx
    report["special_rows"] = special_rows

    # Validation + Formula audit FAIL counts
    def _fail_count(sheet: str, status_col: str) -> int:
        if sheet not in wb.sheetnames:
            return -1
        ws = wb[sheet]
        hh = [str(c.value) if c.value is not None else "" for c in ws[1]]
        si = _col_index(hh, status_col)
        if si is None:
            return -1
        return sum(
            1
            for row in ws.iter_rows(min_row=2)
            if str(row[si - 1].value or "") == "FAIL"
        )

    report["validation_fail_count"] = _fail_count("Validation_Report", "status")
    report["formula_audit_fail_count"] = _fail_count("Formula_Audit", "status")

    # Match_Evidence reviewer/manual fields must be blank.
    evidence_manual_nonblank = 0
    for name in ("reviewer_comment", "manual_status"):
        idx = _col_index(headers, name)
        if idx is None:
            continue
        evidence_manual_nonblank += sum(
            1 for row in data_rows if str(row[idx - 1].value or "").strip()
        )
    report["match_evidence_manual_fields_nonblank"] = evidence_manual_nonblank

    if "Review_Queue" in wb.sheetnames:
        ws = wb["Review_Queue"]
        report["review_queue_count"] = max(0, ws.max_row - 1)
        rq_headers = [str(c.value) if c.value is not None else "" for c in ws[1]]
        report["review_queue_columns_ok"] = all(
            col in rq_headers for col in REVIEW_REQUIRED_COLUMNS + REVIEW_MANUAL_COLUMNS
        )
        report["review_queue_has_priority"] = "priority" in rq_headers
        nonblank = 0
        for name in REVIEW_MANUAL_COLUMNS:
            idx = _col_index(rq_headers, name)
            if idx is None:
                continue
            nonblank += sum(
                1 for row in ws.iter_rows(min_row=2) if str(row[idx - 1].value or "").strip()
            )
        report["review_queue_manual_fields_nonblank"] = nonblank
        priorities = set()
        p_idx = _col_index(rq_headers, "priority")
        if p_idx:
            for row in ws.iter_rows(min_row=2):
                v = str(row[p_idx - 1].value or "").strip()
                if v:
                    priorities.add(v)
        report["review_queue_priorities"] = sorted(priorities)

    # Freeze panes / filters on key sheets.
    key_sheets = ["Match_Evidence", "Summary"]
    freeze_ok = True
    filter_ok = True
    for s in key_sheets:
        if s in wb.sheetnames:
            ws = wb[s]
            if not ws.freeze_panes:
                freeze_ok = False
    if "Match_Evidence" in wb.sheetnames and not wb["Match_Evidence"].auto_filter.ref:
        filter_ok = False
    report["freeze_panes_on_key_sheets"] = freeze_ok
    report["filters_on_match_evidence"] = filter_ok

    # Source-row traceability columns.
    report["traceability_columns_present"] = (
        _col_index(headers, "org_row_id") is not None
        and _col_index(headers, "party_row_id") is not None
    )

    # Central copies present.
    central = path.parents[3] / "04_outputs" / "reconciliation_workbooks" / path.name.replace(
        "final_recon_submission__", "recon_workbook__"
    )
    report["central_recon_copy_exists"] = central.exists() if "02_work_pairs" in str(path) else None

    wb.close()
    return report


def main() -> None:
    parser = argparse.ArgumentParser(description="Inspect recon workbook (read-only).")
    parser.add_argument("paths", nargs="+")
    args = parser.parse_args()
    for p in args.paths:
        report = inspect(Path(p))
        print("=" * 70)
        for key, value in report.items():
            print(f"{key}: {value}")


if __name__ == "__main__":
    main()
