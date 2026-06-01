"""Tests for unknown-label AI grouping and the deterministic review annexure."""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.config import (
    FORMALIZATION_AI_GROUP,
    FORMALIZATION_AI_OFF,
    settings,
)
from src.formalization import ai_label_grouping as grp
from src.formalization.ledger_models import LedgerRow, RecordKind
from src.reconciliation.annexure_builder import build_annexure_plans
from src.reconciliation.recon_models import ReconRow
from src.reconciliation.summary_builder import build_summary_layout
from src.reconciliation.unknown_label_builder import (
    build_unknown_label_groups,
    has_unknown_rows,
)
from src.reconciliation.workbook_writer import write_reconciliation_workbook
from src.reconciliation.working_ledger_builder import build_working_rows


# --------------------------------------------------------------------------- #
# AI grouping module (bounded, off-by-default, no secret, predefined-only)
# --------------------------------------------------------------------------- #
def _unknown_row(raw_type: str, row_no: int) -> LedgerRow:
    return LedgerRow(
        row_id=f"u{row_no}",
        record_kind=RecordKind.TRANSACTION,
        ledger_role="party_ledger",
        detected_ledger_name="Party",
        source_file="p.pdf",
        page_number=1,
        source_row_number=row_no,
        raw_type=raw_type,
        type_label="Unknown",
        particulars="some narration",
        debit_source=100.0,
        raw_text=f"row {row_no} {raw_type}",
        review_flag=True,
        review_reason=f"Type '{raw_type}' unmatched (best score 0 < 70).",
    )


def test_grouping_off_by_default(monkeypatch) -> None:
    monkeypatch.setattr(settings, "ai_formalization_mode", FORMALIZATION_AI_OFF)

    def _boom(*a, **k):  # pragma: no cover - must never run
        raise AssertionError("AI client must not be built when grouping is off")

    monkeypatch.setattr("src.providers.ai_client.build_ai_client", _boom)
    result = grp.group_unknown_labels_batch([_unknown_row("WeirdType", 1)])
    assert result.used_ai is False
    assert result.applied_count == 0


def test_select_unknown_rows_only_transactions() -> None:
    tx = _unknown_row("Foo", 1)
    special = LedgerRow(
        row_id="s1",
        record_kind=RecordKind.OPENING_BALANCE,
        ledger_role="party_ledger",
        detected_ledger_name="Party",
        source_file="p.pdf",
        page_number=1,
        source_row_number=2,
        type_label="OpeningBalance",
    )
    known = _unknown_row("Bar", 3)
    known.type_label = "Invoice"
    selected = grp.select_unknown_rows([tx, special, known])
    assert [r.row_id for r in selected] == ["u1"]


def test_evidence_is_bounded_and_lists_predefined_labels() -> None:
    rows = [_unknown_row("Foo", i) for i in range(10)]
    groups = grp._group_unknown_rows(rows)
    labels = grp.predefined_labels()
    evidence = grp.build_grouping_evidence(groups, labels)
    assert evidence["predefined_labels"] == labels
    # Distinct raw type "Foo" -> one group, samples capped.
    assert len(evidence["unknown_groups"]) == 1
    assert len(evidence["unknown_groups"][0]["samples"]) <= grp._SAMPLE_ROWS_PER_GROUP
    assert evidence["unknown_groups"][0]["affected_row_count"] == 10
    # Payload must respect the character budget.
    text = json.dumps(evidence, ensure_ascii=False)
    assert len(text) <= settings.ai_max_input_chars_per_request


def test_validate_suggestion_predefined_only() -> None:
    allowed = set(grp.predefined_labels())
    # Confident predefined label is accepted.
    label, conf, _ = grp._validate_suggestion(
        {"suggested_label": "Invoice", "confidence": 95, "reason": "sale"}, allowed
    )
    assert label == "Invoice"
    assert conf == 95
    # Invented label is rejected to the review sentinel.
    label, _, _ = grp._validate_suggestion(
        {"suggested_label": "MadeUpLabel", "confidence": 99}, allowed
    )
    assert label == grp.UNKNOWN_NEEDS_REVIEW
    # Low-confidence predefined label is rejected to the review sentinel.
    label, _, _ = grp._validate_suggestion(
        {"suggested_label": "Invoice", "confidence": 10}, allowed
    )
    assert label == grp.UNKNOWN_NEEDS_REVIEW


def test_apply_suggestion_changes_label_not_raw_type() -> None:
    row = _unknown_row("SomeSale", 1)
    grp._apply_suggestion(row, "Invoice", 90.0, "looks like a sale")
    assert row.type_label == "Invoice"
    assert row.raw_type == "SomeSale"  # raw never altered
    assert row.ai_label_applied is True
    assert "unmatched" not in row.review_reason.lower()


def test_cache_hit_applies_without_network(monkeypatch, tmp_path: Path) -> None:
    monkeypatch.setattr(settings, "ai_formalization_mode", FORMALIZATION_AI_GROUP)
    monkeypatch.setattr(settings, "ai_formalization_cache_enabled", True)
    monkeypatch.setattr(settings, "ai_formalization_cache_dir", tmp_path)

    def _boom(*a, **k):  # pragma: no cover - cache hit must avoid network
        raise AssertionError("network must not be called on cache hit")

    monkeypatch.setattr("src.providers.ai_client.build_ai_client", _boom)

    rows = [_unknown_row("BankPmt", 1)]
    groups = grp._group_unknown_rows(grp.select_unknown_rows(rows))
    payload = grp.build_grouping_evidence(groups, grp.predefined_labels())
    cache_key = grp._cache_key(payload)
    (tmp_path / f"{cache_key}.json").write_text(
        json.dumps(
            {"groups": [{"raw_type": "BankPmt", "suggested_label": "Payment",
                         "confidence": 92, "reason": "bank payment"}]}
        ),
        encoding="utf-8",
    )
    result = grp.group_unknown_labels_batch(rows)
    assert result.cache_hit is True
    assert result.applied_count == 1
    assert rows[0].type_label == "Payment"


# --------------------------------------------------------------------------- #
# Deterministic grouping report + review annexure presence
# --------------------------------------------------------------------------- #
def _recon_unknown(row_id: str, role: str, raw_type: str) -> ReconRow:
    return ReconRow(
        data={
            "row_id": row_id,
            "record_kind": "Transaction",
            "ledger_role": role,
            "type_label": "Unknown",
            "raw_type": raw_type,
            "particulars": "mystery line",
            "reference_no": f"R-{row_id}",
            "normalized_reference": f"r{row_id}",
            "debit_source": 50.0,
            "credit_source": 0.0,
            "raw_text": "mystery line raw",
            "review_flag": True,
        },
        sheet_name=role,
    )


def test_build_unknown_groups_counts_sides() -> None:
    org = [_recon_unknown("o1", "org_ledger", "Mystery")]
    party = [
        _recon_unknown("p1", "party_ledger", "Mystery"),
        _recon_unknown("p2", "party_ledger", "Mystery"),
    ]
    assert has_unknown_rows(org, party) is True
    groups = build_unknown_label_groups(org, party)
    assert len(groups) == 1
    g = groups[0]
    assert g["affected_row_count"] == 3
    assert g["org_row_count"] == 1
    assert g["party_row_count"] == 2
    assert g["review_required"] is True


def _known(row_id: str, role: str) -> ReconRow:
    return ReconRow(
        data={
            "row_id": row_id,
            "record_kind": "Transaction",
            "ledger_role": role,
            "type_label": "Invoice",
            "raw_type": "Sale",
            "reference_no": f"R-{row_id}",
            "normalized_reference": f"r{row_id}",
            "debit_source": 100.0,
            "credit_source": 0.0,
            "raw_text": "sale",
            "review_flag": False,
        },
        sheet_name=role,
    )


def _write(tmp_path: Path, org_rows, party_rows) -> Path:
    labels = ["Invoice", "Payment"]
    annex = build_annexure_plans(labels)
    summary = build_summary_layout(
        labels, annex_sheet_by_label={p.label: p.sheet_name for p in annex}
    )
    out = tmp_path / "recon.xlsx"
    write_reconciliation_workbook(
        output_path=out,
        pair_id="pair_x",
        formalized_path=tmp_path / "f.xlsx",
        reference_path="ref.xlsx",
        labels=labels,
        org_sheet_name="ORG",
        party_sheet_name="PARTY",
        org_rows=org_rows,
        party_rows=party_rows,
        org_working=build_working_rows(org_rows),
        party_working=build_working_rows(party_rows),
        summary_layout=summary,
        annex_plans=annex,
        matches=[],
        validation_items=[],
    )
    return out


def test_unknown_annexure_present_only_when_needed(tmp_path: Path) -> None:
    # With an unknown row -> review annexure exists; canonical annex count unchanged.
    out = _write(
        tmp_path,
        [_known("o1", "org_ledger")],
        [_recon_unknown("p1", "party_ledger", "Mystery")],
    )
    wb = load_workbook(out)
    assert "Unknown_Needs_Review" in wb.sheetnames
    annex = [s for s in wb.sheetnames if s.startswith("Annex_")]
    assert len(annex) == 2
    text = {str(c.value) for row in wb["Unknown_Needs_Review"].iter_rows()
            for c in row if c.value is not None}
    assert any("Section A" in t for t in text)
    assert any("Mystery" in t for t in text)
    wb.close()


def test_unknown_annexure_absent_when_all_known(tmp_path: Path) -> None:
    out = _write(
        tmp_path,
        [_known("o1", "org_ledger")],
        [_known("p1", "party_ledger")],
    )
    wb = load_workbook(out)
    assert "Unknown_Needs_Review" not in wb.sheetnames
    wb.close()
