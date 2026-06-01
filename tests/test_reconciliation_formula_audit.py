from __future__ import annotations

from openpyxl import Workbook

from src.reconciliation.formula_builder import (
    audit_formula_cells,
    audit_protected_cells,
)


def test_formula_audit_catches_missing_formula() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "=SUM(1,2)"
    ws["A2"] = 3
    audit = audit_formula_cells(ws, ["A1", "A2"], "summary")
    statuses = {a.cell: a.status for a in audit}
    assert statuses["A1"] == "PASS"
    assert statuses["A2"] == "FAIL"


def test_formula_audit_blank_cell_fails() -> None:
    wb = Workbook()
    ws = wb.active
    audit = audit_formula_cells(ws, ["B5"], "summary")
    assert audit[0].status == "FAIL"


def test_formula_audit_detects_wrong_column_reference() -> None:
    # amount_difference must reference org_amount (P) and party_amount (Q).
    wb = Workbook()
    ws = wb.active
    ws.title = "Master_Match_Table"
    ws["R2"] = "=ABS(ABS(Q2)-ABS(R2))"  # wrong: references Q and R, not P and Q
    audit = audit_formula_cells(ws, ["R2"], "match_amount_difference", expected_contains=["P", "Q"])
    assert audit[0].status == "FAIL"
    assert "expected column" in audit[0].issue.lower()


def test_formula_audit_accepts_correct_reference() -> None:
    wb = Workbook()
    ws = wb.active
    ws["R2"] = '=IF(OR(P2="",Q2=""),"",ABS(ABS(P2)-ABS(Q2)))'
    audit = audit_formula_cells(ws, ["R2"], "match_amount_difference", expected_contains=["P", "Q"])
    assert audit[0].status == "PASS"


def test_protected_cell_with_formula_fails() -> None:
    # Evidence columns (e.g. org_raw_type) must never contain a formula.
    wb = Workbook()
    ws = wb.active
    ws["S2"] = "=ABS(P2-Q2)"  # corruption: formula written into evidence column
    ws["S3"] = "Sale"
    audit = audit_protected_cells(ws, ["S2", "S3"], "match_protected_evidence")
    statuses = {a.cell: a.status for a in audit}
    assert statuses["S2"] == "FAIL"
    assert statuses["S3"] == "PASS"

