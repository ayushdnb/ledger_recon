"""Tests for the submission-quality presentation layer and safety invariants."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from src.reconciliation.ai_availability import (
    _PROBE_SYSTEM,
    _PROBE_USER,
    check_ai_availability,
)
from src.reconciliation.annexure_builder import build_annexure_plans
from src.reconciliation.recon_models import MatchRecord, ReconRow, ValidationItem
from src.reconciliation.review_queue_builder import (
    MANUAL_COLUMNS,
    REVIEW_QUEUE_COLUMNS,
    build_review_queue,
)
from src.reconciliation.summary_builder import build_summary_layout
from src.reconciliation.team_workbook_builder import build_team_workbook
from src.reconciliation.workbook_writer import write_reconciliation_workbook
from src.reconciliation.working_ledger_builder import build_working_rows

LABELS = ["OpeningBalance", "Invoice", "Payment", "ClosingBalance"]


def _is_formula(value: object) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _row(row_id: str, role: str, sheet: str, label: str = "Invoice", *, ref: str | None = None) -> ReconRow:
    return ReconRow(
        data={
            "row_id": row_id,
            "record_kind": "Transaction",
            "ledger_role": role,
            "detected_ledger_name": sheet,
            "date": "2025-01-01",
            "normalized_date": "2025-01-01",
            "raw_type": label,
            "type_label": label,
            "reference_no": ref if ref is not None else f"REF-{row_id}",
            "normalized_reference": (ref if ref is not None else f"ref{row_id}").lower(),
            "debit_source": 100.0,
            "credit_source": 0.0,
            "debit_org_perspective": 100.0,
            "credit_org_perspective": 0.0,
            "raw_text": "sample",
            "review_flag": False,
            "confidence_score": 99.0,
        },
        sheet_name=sheet,
    )


def _build(tmp_path: Path, *, matches=None, annex_labels=None):
    labels = annex_labels or LABELS
    org_rows = [_row("o1", "org_ledger", "ORG")]
    party_rows = [_row("p1", "party_ledger", "PARTY")]
    if matches is None:
        matches = [
            MatchRecord(
                match_group_id="MG000001",
                match_status="matched_strong",
                match_confidence=1.0,
                match_rule="stage1_exact_ref_type_amount",
                type_label="Invoice",
                org_row_id="o1",
                party_row_id="p1",
                org_normalized_reference="refo1",
                party_normalized_reference="refo1",
                org_amount=100.0,
                party_amount=100.0,
                amount_difference=0.0,
            )
        ]
    annex = build_annexure_plans(labels)
    summary = build_summary_layout(
        labels, annex_sheet_by_label={p.label: p.sheet_name for p in annex}
    )
    output = tmp_path / "recon.xlsx"
    write_reconciliation_workbook(
        output_path=output,
        pair_id="pair_x",
        formalized_path=tmp_path / "formalized.xlsx",
        reference_path="ref.xlsx",
        labels=labels,
        org_sheet_name="ORG",
        party_sheet_name="PARTY",
        org_rows=org_rows,
        party_rows=party_rows,
        org_working=build_working_rows(org_rows),
        party_working=build_working_rows(party_rows),
        summary_layout=summary,
        annex_plans=annex,
        matches=matches,
        validation_items=[ValidationItem("x", "s", "PASS", 1, "")],
        recon_period="Jan'25 to Jun'25",
        org_ledger_name="Org Co",
        party_ledger_name="Party Co",
    )
    return output


def test_required_sheets_and_order(tmp_path: Path) -> None:
    wb = load_workbook(_build(tmp_path), data_only=False)
    names = wb.sheetnames
    assert names[0] == "README"
    assert names[1] == "Executive_Summary"
    for required in ("Summary", "AI_Decision_Audit", "Match_Evidence", "Review_Queue",
                     "Validation_Report", "Formula_Audit", "Assumptions_And_Limits"):
        assert required in names
    # Executive_Summary must come before Summary; Assumptions last block.
    assert names.index("Executive_Summary") < names.index("Summary")
    wb.close()


def test_executive_summary_sections(tmp_path: Path) -> None:
    wb = load_workbook(_build(tmp_path), data_only=False)
    ws = wb["Executive_Summary"]
    text = {str(c.value) for row in ws.iter_rows() for c in row if c.value is not None}
    for section in ("Workbook Identity", "Reconciliation Status Snapshot",
                    "Amount Snapshot", "Submission Note"):
        assert section in text
    # Snapshot counts are formula-driven.
    formulas = [c.value for row in ws.iter_rows() for c in row if _is_formula(c.value)]
    assert any("COUNTIF" in f for f in formulas)
    wb.close()


def test_summary_top_block_and_formula_amounts(tmp_path: Path) -> None:
    wb = load_workbook(_build(tmp_path), data_only=False)
    ws = wb["Summary"]
    text = {str(c.value) for row in ws.iter_rows() for c in row if c.value is not None}
    assert "Vendor / Party Name" in text
    assert "Recon Period" in text
    assert "Party Co" in text
    # Label-row org/party amounts must be formulas (find the Invoice row).
    invoice_row = None
    for row in ws.iter_rows():
        for c in row:
            if c.value == "Invoice":
                invoice_row = c.row
    assert invoice_row is not None
    assert _is_formula(ws[f"C{invoice_row}"].value)
    assert _is_formula(ws[f"D{invoice_row}"].value)
    assert _is_formula(ws[f"E{invoice_row}"].value)
    wb.close()


def test_summary_has_annexure_hyperlinks(tmp_path: Path) -> None:
    wb = load_workbook(_build(tmp_path), data_only=False)
    ws = wb["Summary"]
    links = [c.hyperlink.target for row in ws.iter_rows() for c in row if c.hyperlink]
    assert any("Annex_" in (t or "") for t in links)
    wb.close()


def test_annexures_for_all_labels_and_complete(tmp_path: Path) -> None:
    wb = load_workbook(_build(tmp_path), data_only=False)
    annex_sheets = [s for s in wb.sheetnames if s.startswith("Annex_")]
    assert len(annex_sheets) == len(LABELS)
    # An empty annexure (Payment has no matching tx) still has all four sections.
    ws = wb["Annex_Payment"]
    text = {str(c.value) for row in ws.iter_rows() for c in row if c.value is not None}
    assert any("Section A" in t for t in text)
    assert any("Section B" in t for t in text)
    assert any("Section C" in t for t in text)
    assert any("Section D" in t for t in text)
    assert any("No transactions found" in t for t in text)
    wb.close()


def test_review_queue_columns_priority_and_blank_manuals(tmp_path: Path) -> None:
    # Build with an unmatched + ambiguous match to populate the queue.
    matches = [
        MatchRecord(
            match_group_id="MG1", match_status="unmatched_org", match_confidence=0.0,
            match_rule="stage7_unmatched_org", type_label="Invoice", org_row_id="o1",
            org_normalized_reference="refo1", org_amount=100.0, review_required=True,
            review_reason="No deterministic match found for org row.",
        ),
        MatchRecord(
            match_group_id="MG2", match_status="candidate_review", match_confidence=0.7,
            match_rule="stage1_2_ambiguous_exact_ref", type_label="Invoice", org_row_id="o1",
            party_row_id="p1", org_normalized_reference="refo1", party_normalized_reference="refo1",
            org_amount=100.0, party_amount=100.0, amount_difference=0.0, review_required=True,
            review_reason="Multiple exact-reference candidates; deterministic tie unresolved.",
        ),
    ]
    wb = load_workbook(_build(tmp_path, matches=matches), data_only=False)
    ws = wb["Review_Queue"]
    headers = [str(c.value) for c in ws[1]]
    assert headers == REVIEW_QUEUE_COLUMNS
    assert "priority" in headers
    # Manual columns must be blank in every data row.
    for name in MANUAL_COLUMNS:
        idx = headers.index(name)
        for row in ws.iter_rows(min_row=2):
            assert not str(row[idx].value or "").strip()
    # At least one HIGH priority (the ambiguous candidate).
    p_idx = headers.index("priority")
    priorities = {str(row[p_idx].value) for row in ws.iter_rows(min_row=2)}
    assert "HIGH" in priorities
    wb.close()


def test_assumptions_sheet_present(tmp_path: Path) -> None:
    wb = load_workbook(_build(tmp_path), data_only=False)
    ws = wb["Assumptions_And_Limits"]
    text = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
    assert "deterministically post-validated" in text
    assert "source of truth" in text.lower()
    wb.close()


def test_styling_basics_applied(tmp_path: Path) -> None:
    wb = load_workbook(_build(tmp_path), data_only=False)
    mmt = wb["Match_Evidence"]
    assert mmt.freeze_panes  # frozen header
    assert mmt.auto_filter.ref  # filter applied
    assert mmt["A1"].font.bold  # bold header
    # amount_difference column carries the amount number format.
    headers = [str(c.value) for c in mmt[1]]
    diff_col = get_column_letter(headers.index("amount_difference") + 1)
    assert "#,##0" in (mmt[f"{diff_col}2"].number_format or "")
    wb.close()


def test_no_formulas_in_evidence_columns(tmp_path: Path) -> None:
    wb = load_workbook(_build(tmp_path), data_only=False)
    ws = wb["Match_Evidence"]
    headers = [str(c.value) for c in ws[1]]
    for name in ("org_raw_type", "party_raw_type", "org_particulars", "party_particulars",
                 "org_reference", "party_reference", "org_normalized_reference",
                 "party_normalized_reference"):
        idx = headers.index(name)
        for row in ws.iter_rows(min_row=2):
            assert not _is_formula(row[idx].value)
    wb.close()


def test_team_workbook_is_clean_and_uses_excel_tables(tmp_path: Path) -> None:
    matches = [
        MatchRecord(
            match_group_id="AI-1",
            match_status="matched_ai",
            match_confidence=0.95,
            match_rule="ai_arbitration_amount_date",
            type_label="Invoice",
            org_row_id="o1",
            party_row_id="p1",
            org_amount=100.0,
            party_amount=100.0,
            amount_difference=0.0,
            decision_source="AI",
            validation_status="ACCEPTED",
        )
    ]
    internal = _build(tmp_path, matches=matches)
    team = tmp_path / "team.xlsx"
    build_team_workbook(
        internal_path=internal,
        output_path=team,
        pair_id="pair_x",
        raw_sheet_names=["ORG", "PARTY"],
    )
    wb = load_workbook(team, data_only=False)
    assert wb.sheetnames[0] == "Team_Guide"
    for name in ("ORG", "PARTY", "Match_Evidence", "AI_Decision_Audit", "Review_Queue"):
        assert wb[name].sheet_state == "hidden"
    for name in ("Summary", "Annex_Invoice", "Rows_Needing_Improvement", "Working ORG"):
        assert wb[name].sheet_state == "visible"
    working = wb["Working ORG"]
    assert working.tables
    headers = [str(cell.value) for cell in working[1]]
    decision_col = get_column_letter(headers.index("decision_source") + 1)
    assert working.column_dimensions[decision_col].hidden is True
    annex_text = {
        str(cell.value)
        for row in wb["Annex_Invoice"].iter_rows()
        for cell in row
        if cell.value is not None
    }
    assert "matched_ai" not in annex_text
    assert "AI" not in annex_text
    wb.close()


# --- AI availability ------------------------------------------------------- #
def test_ai_probe_contains_no_financial_data() -> None:
    blob = (_PROBE_SYSTEM + " " + _PROBE_USER).lower()
    for forbidden in ("ledger", "invoice", "amount", "debit", "credit", "reference",
                      "vendor", "party"):
        assert forbidden not in blob
    # No digits beyond the literal probe JSON.
    assert "100" not in blob


def test_ai_disabled_reports_unavailable() -> None:
    from src.config import settings

    cfg = settings.model_copy(update={"ai_enabled": False})
    result = check_ai_availability(cfg)
    assert result.available is False
    assert result.attempted is False


def test_ai_probe_sends_only_probe_messages(monkeypatch) -> None:
    from src.config import settings
    import src.providers.ai_client as ai_client

    captured: list[list[dict]] = []

    class _Fake:
        def complete_chat(self, messages, *, request_json_object=True):
            captured.append(messages)
            return '{"ok": true}'

    monkeypatch.setattr(ai_client, "build_ai_client", lambda config=None: _Fake())
    # Local provider => approved without hosted gate.
    cfg = settings.model_copy(update={"ai_enabled": True, "ai_provider": "local_probe"})
    result = check_ai_availability(cfg)
    assert result.available is True
    assert captured, "probe must have sent messages"
    text = " ".join(m["content"] for m in captured[0]).lower()
    for forbidden in ("ledger", "invoice", "debit", "credit", "amount"):
        assert forbidden not in text
