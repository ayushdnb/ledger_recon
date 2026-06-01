from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from src.reconciliation.reference_profile import profile_reference_workbook


def test_reference_profile_reads_sheet_structure(tmp_path: Path) -> None:
    path = tmp_path / "reference.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Header"
    ws["A2"] = "=SUM(1,2)"
    wb.save(path)
    wb.close()

    profile = profile_reference_workbook(path)
    assert profile["sheet_count"] == 1
    assert "Summary" in profile["sheet_names"]
    assert profile["sheets"][0]["formula_count"] >= 1

