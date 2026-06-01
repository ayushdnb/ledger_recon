from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from src.reconciliation.annexure_builder import build_annexure_plans
from src.reconciliation.matching_engine import is_missing_reference, reconcile_rows
from src.reconciliation.recon_models import MatchRecord, ReconRow, ValidationItem
from src.reconciliation.summary_builder import build_summary_layout
from src.reconciliation.workbook_writer import write_reconciliation_workbook
from src.reconciliation.working_ledger_builder import build_working_rows


def _col(headers: list[str], name: str) -> str:
    return get_column_letter([str(h) for h in headers].index(name) + 1)


def _is_formula(value: object) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _row(row_id: str, role: str, sheet: str, label: str = "Invoice") -> ReconRow:
    return ReconRow(
        data={
            "row_id": row_id,
            "record_kind": "Transaction",
            "ledger_role": role,
            "detected_ledger_name": sheet,
            "source_file": "x.pdf",
            "page_number": 1,
            "source_row_number": 1,
            "date": "2025-01-01",
            "normalized_date": "2025-01-01",
            "raw_type": label,
            "type_label": label,
            "reference_no": f"REF-{row_id}",
            "normalized_reference": f"ref{row_id}".lower(),
            "debit_source": 100.0,
            "credit_source": 0.0,
            "debit_org_perspective": 100.0,
            "credit_org_perspective": 0.0,
            "raw_text": "sample",
            "review_flag": False,
            "review_reason": "",
            "confidence_score": 99.0,
        },
        sheet_name=sheet,
    )


def test_final_workbook_creation_and_structure(tmp_path: Path) -> None:
    org_rows = [_row("o1", "org_ledger", "ORG")]
    party_rows = [_row("p1", "party_ledger", "PARTY")]
    org_working = build_working_rows(org_rows)
    party_working = build_working_rows(party_rows)
    labels = ["Invoice", "Payment"]
    summary = build_summary_layout(labels)
    annex = build_annexure_plans(labels)
    matches = [
        MatchRecord(
            match_group_id="MG000001",
            match_status="matched_strong",
            match_confidence=1.0,
            match_rule="stage1_exact_ref_type_amount",
            type_label="Invoice",
            org_row_id="o1",
            party_row_id="p1",
            org_amount=100.0,
            party_amount=100.0,
            amount_difference=0.0,
        )
    ]
    output = tmp_path / "recon.xlsx"
    write_reconciliation_workbook(
        output_path=output,
        pair_id="pair_x",
        formalized_path=tmp_path / "formalized.xlsx",
        reference_path="",
        labels=labels,
        org_sheet_name="ORG",
        party_sheet_name="PARTY",
        org_rows=org_rows,
        party_rows=party_rows,
        org_working=org_working,
        party_working=party_working,
        summary_layout=summary,
        annex_plans=annex,
        matches=matches,
        validation_items=[ValidationItem("x", "s", "PASS", 1, "")],
    )
    assert output.exists()

    wb = load_workbook(output, data_only=False)
    assert wb.sheetnames[0] == "README"
    assert "Summary" in wb.sheetnames
    assert "AI_Decision_Audit" in wb.sheetnames
    assert "Match_Evidence" in wb.sheetnames
    assert "Master_Match_Table" not in wb.sheetnames
    assert "Formula_Audit" in wb.sheetnames
    assert any(s.startswith("Annex_") for s in wb.sheetnames)
    ws_work = wb["Working ORG"]
    assert isinstance(ws_work["N2"].value, str) and ws_work["N2"].value.startswith("=")
    assert isinstance(ws_work["Q2"].value, str) and ws_work["Q2"].value.startswith("=")
    wb.close()


def _row_with_raw_type(row_id: str, role: str, sheet: str, raw_type: str) -> ReconRow:
    r = _row(row_id, role, sheet)
    r.data["raw_type"] = raw_type
    return r


def test_amount_difference_formula_in_correct_column_and_evidence_preserved(tmp_path: Path) -> None:
    org_rows = [_row_with_raw_type("o1", "org_ledger", "ORG", "Sale")]
    party_rows = [_row_with_raw_type("p1", "party_ledger", "PARTY", "Purchase")]
    labels = ["Invoice", "Payment"]
    matches = [
        MatchRecord(
            match_group_id="MG000001",
            match_status="matched_strong",
            match_confidence=1.0,
            match_rule="stage1_exact_ref_type_amount",
            type_label="Invoice",
            org_row_id="o1",
            party_row_id="p1",
            org_normalized_reference="refo1",
            party_normalized_reference="refp1",
            org_amount=100.0,
            party_amount=100.0,
            amount_difference=0.0,
            org_raw_type="Sale",
            party_raw_type="Purchase",
        )
    ]
    output = tmp_path / "recon.xlsx"
    write_reconciliation_workbook(
        output_path=output,
        pair_id="pair_x",
        formalized_path=tmp_path / "formalized.xlsx",
        reference_path="",
        labels=labels,
        org_sheet_name="ORG",
        party_sheet_name="PARTY",
        org_rows=org_rows,
        party_rows=party_rows,
        org_working=build_working_rows(org_rows),
        party_working=build_working_rows(party_rows),
        summary_layout=build_summary_layout(labels),
        annex_plans=build_annexure_plans(labels),
        matches=matches,
        validation_items=[ValidationItem("x", "s", "PASS", 1, "")],
    )

    wb = load_workbook(output, data_only=False)
    ws = wb["Match_Evidence"]
    headers = [c.value for c in ws[1]]
    diff_col = _col(headers, "amount_difference")
    org_type_col = _col(headers, "org_raw_type")
    party_type_col = _col(headers, "party_raw_type")
    org_amt_col = _col(headers, "org_amount")
    party_amt_col = _col(headers, "party_amount")

    diff_value = ws[f"{diff_col}2"].value
    assert _is_formula(diff_value), "amount_difference must be a formula"
    assert org_amt_col in diff_value and party_amt_col in diff_value

    # Evidence columns must keep their source text and never be a formula.
    assert ws[f"{org_type_col}2"].value == "Sale"
    assert ws[f"{party_type_col}2"].value == "Purchase"
    assert not _is_formula(ws[f"{org_type_col}2"].value)
    assert not _is_formula(ws[f"{party_type_col}2"].value)

    # Formula_Audit must not contain any FAIL.
    ws_fa = wb["Formula_Audit"]
    fa_headers = [c.value for c in ws_fa[1]]
    status_idx = [str(h) for h in fa_headers].index("status")
    fails = [
        row[status_idx]
        for row in ws_fa.iter_rows(min_row=2, values_only=True)
        if row[status_idx] == "FAIL"
    ]
    assert fails == [], f"Formula_Audit reported FAIL rows: {len(fails)}"
    wb.close()


def _formalized_paths() -> list[tuple[str, Path]]:
    base = Path(__file__).resolve().parents[1] / "data" / "02_work_pairs"
    out: list[tuple[str, Path]] = []
    for pair in ("pair_001_baby_and_mom__good_luck", "pair_002_elegant_crafts_india"):
        p = base / pair / "output" / f"formalized_ledgers__{pair}.xlsx"
        out.append((pair, p))
    return out


@pytest.mark.parametrize("pair_id,formalized", _formalized_paths())
def test_recon_generation_for_real_pairs(pair_id: str, formalized: Path, tmp_path: Path) -> None:
    if not formalized.exists():
        pytest.skip(f"formalized workbook not present for {pair_id}")
    from src.reconciliation.ledger_loader import load_formalized_ledgers

    loaded = load_formalized_ledgers(formalized)
    labels = [
        "Invoice", "CreditNote", "DebitNote", "Payment", "Receipt",
        "TDS", "PQV", "Journal", "OpeningBalance", "ClosingBalance",
    ]
    org_working = build_working_rows(loaded.org_rows)
    party_working = build_working_rows(loaded.party_rows)
    result = reconcile_rows(loaded.org_rows, loaded.party_rows)

    # No strong match may rest on a missing reference.
    strong_missing = [
        m
        for m in result.records
        if m.match_status == "matched_strong"
        and (
            is_missing_reference(m.org_normalized_reference)
            or is_missing_reference(m.party_normalized_reference)
        )
    ]
    assert strong_missing == [], f"{pair_id}: {len(strong_missing)} strong matches with missing ref"

    output = tmp_path / f"recon_{pair_id}.xlsx"
    formula_audit = write_reconciliation_workbook(
        output_path=output,
        pair_id=pair_id,
        formalized_path=formalized,
        reference_path="",
        labels=labels,
        org_sheet_name=loaded.org_sheet,
        party_sheet_name=loaded.party_sheet,
        org_rows=loaded.org_rows,
        party_rows=loaded.party_rows,
        org_working=org_working,
        party_working=party_working,
        summary_layout=build_summary_layout(labels),
        annex_plans=build_annexure_plans(labels),
        matches=result.records,
        validation_items=[],
    )
    assert output.exists()
    assert sum(1 for f in formula_audit if f.status == "FAIL") == 0

    wb = load_workbook(output, data_only=False)
    ws = wb["Match_Evidence"]
    headers = [str(c.value) for c in ws[1]]
    org_type_idx = headers.index("org_raw_type")
    party_type_idx = headers.index("party_raw_type")
    # No evidence column cell may be a formula in any data row.
    evidence_formulas = sum(
        1
        for row in ws.iter_rows(min_row=2, values_only=True)
        if _is_formula(row[org_type_idx]) or _is_formula(row[party_type_idx])
    )
    assert evidence_formulas == 0
    wb.close()
